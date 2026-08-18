"""
AMS System Proxy Server (NAS版)
Run: python3 proxy-server.py
Access: http://localhost:8899/

Multi-threaded server for concurrent API requests.
"""
from http.server import HTTPServer, BaseHTTPRequestHandler
from socketserver import ThreadingMixIn
import urllib.request, json, os, uuid, sqlite3, hashlib, secrets
from pathlib import Path

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

# 配置
TOKEN_PATH = os.path.join(BASE_DIR, "cache", "ams_token.txt")
AUTH_DB = os.path.join(BASE_DIR, "auth.db")
AMS_BASE = "https://ams.xintujing.online"
API_BASE = "https://api.xintujing.online"
HTML_PATH = os.path.join(BASE_DIR, "ams_system.html")
PORT = 8899
SESSION_DAYS = 7

# 校区ID映射
HOTEL_IDS = {
    "10144": "上岸公寓B座",
    "10137": "上岸公寓C座",
    "10143": "上岸公寓D座",
    "10145": "小新公寓",
    "10148": "景然力沃校区",
    "10138": "嵘泰校区",
    "10139": "塔利北校区",
    "10140": "塔利南校区",
    "10163": "城际酒店",
}

# ===== 账号权限系统 =====
# 模块目录（key 与前端页面一致）
MODULES = {
    "summary": "数据汇总",
    "info": "基本信息",
    "weekly": "周效比",
    "staff": "花名册",
    "residents": "在住名单",
    "rooms": "楼栋房间",
    "income": "收入报表",
    "finance": "财务数据",
    "payroll": "工资表",
    "linen": "布草管理",
    "board": "包住对比",
    "test": "校区地图",
    "bedmanage": "床位管理",
}

# 默认工作区（首次启动自动创建，可在【权限管理】中调整）
DEFAULT_WORKSPACES = [
    {"name": "管理后台", "modules": list(MODULES.keys()), "campuses": [], "remark": "管理员专用，拥有全部模块"},
    {"name": "公寓运营", "modules": ["summary", "info", "weekly", "staff", "residents", "rooms", "test", "bedmanage"],
     "campuses": [], "remark": "公寓一组/二组日常运营"},
    {"name": "财务", "modules": ["summary", "income", "finance", "payroll", "board"],
     "campuses": [], "remark": "收入/财务/工资/包住对比"},
    {"name": "酒店前台", "modules": ["residents", "rooms", "bedmanage", "test"],
     "campuses": [], "remark": "前台接待日常使用"},
]

# 接口 → 模块映射（路径前缀命中即算该模块，管理员不受限）
API_MODULE_RULES = [
    ("/api/rooms/query", ["rooms"]),
    ("/api/room/clean", ["rooms"]),
    ("/api/finance/query", ["finance"]),
    ("/api/occupancy/", ["summary", "residents", "info"]),
    ("/api/residents/", ["residents", "staff"]),
    ("/api/income/summary", ["summary", "income"]),
    ("/api/income/records", ["income"]),
    ("/api/weekly/", ["weekly"]),
    ("/api/board/", ["board"]),
    ("/api/test", ["test"]),
    ("/api/wake/", ["bedmanage"]),
    ("/api/battle/", ["bedmanage"]),
    ("/api/smokefree/", ["bedmanage"]),
    ("/api/roomtypes/", ["bedmanage"]),
    ("/api/linen/", ["linen"]),
    ("/api/payroll/", ["payroll"]),
    ("/api/hotel/web/basics/", ["info", "rooms", "bedmanage", "test"]),
    ("/api/hotel/web/business/checkOrder/", ["staff", "residents", "bedmanage"]),
    ("/api/hotel/web/", ["info", "rooms", "staff", "residents", "bedmanage", "test"]),
]


def _hash_password(password, salt):
    """PBKDF2-SHA256 口令散列（内置库，无需额外依赖）"""
    return hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"),
                               salt.encode("utf-8"), 100000).hex()


def init_auth_db():
    """初始化账号权限库：建表 + 首次启动自动创建默认工作区和管理员"""
    os.makedirs(os.path.dirname(AUTH_DB), exist_ok=True)
    conn = sqlite3.connect(AUTH_DB)
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS workspaces (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT NOT NULL UNIQUE,
        modules TEXT NOT NULL DEFAULT '[]',
        campuses TEXT NOT NULL DEFAULT '[]',
        remark TEXT DEFAULT '',
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT NOT NULL UNIQUE,
        password_hash TEXT NOT NULL,
        salt TEXT NOT NULL,
        display_name TEXT DEFAULT '',
        workspace_ids TEXT NOT NULL DEFAULT '[]',
        is_admin INTEGER NOT NULL DEFAULT 0,
        enabled INTEGER NOT NULL DEFAULT 1,
        created_at TEXT DEFAULT (datetime('now','localtime'))
    )""")
    c.execute("""CREATE TABLE IF NOT EXISTS sessions (
        token TEXT PRIMARY KEY,
        user_id INTEGER NOT NULL,
        created_at TEXT DEFAULT (datetime('now','localtime')),
        expires_at TEXT NOT NULL
    )""")
    if c.execute("SELECT COUNT(*) FROM workspaces").fetchone()[0] == 0:
        for ws in DEFAULT_WORKSPACES:
            c.execute("INSERT INTO workspaces (name, modules, campuses, remark) VALUES (?,?,?,?)",
                      (ws["name"], json.dumps(ws["modules"], ensure_ascii=False),
                       json.dumps(ws["campuses"]), ws.get("remark", "")))
    if c.execute("SELECT COUNT(*) FROM users").fetchone()[0] == 0:
        salt = secrets.token_hex(16)
        pw_hash = _hash_password("admin123", salt)
        c.execute("""INSERT INTO users (username, password_hash, salt, display_name, workspace_ids, is_admin, enabled)
                     VALUES (?,?,?,?,?,1,1)""",
                  ("admin", pw_hash, salt, "系统管理员", "[]"))
        try:
            cache_dir = os.path.join(BASE_DIR, "cache")
            os.makedirs(cache_dir, exist_ok=True)
            with open(os.path.join(cache_dir, "初始管理员密码.txt"), "w", encoding="utf-8") as f:
                f.write("默认管理员账号：admin\n默认密码：admin123\n请登录后尽快在【权限管理】中修改密码。\n")
        except Exception:
            pass
    conn.commit()
    conn.close()

class ThreadedHTTPServer(ThreadingMixIn, HTTPServer):
    """Multi-threaded — REQUIRED: frontend sends parallel requests."""
    daemon_threads = True

class ProxyHandler(BaseHTTPRequestHandler):
    def do_OPTIONS(self):
        self.send_cors(200)
        self.end_headers()

    def do_GET(self):
        from urllib.parse import urlparse
        parsed_path = urlparse(self.path)
        path = parsed_path.path

        # 模块权限校验（/api/status 保持公开，供健康检查）
        if path.startswith("/api/") and path != "/api/status":
            code, msg = self._check_api_access(path)
            if code:
                return self.send_json({"code": code, "msg": msg}, code)

        if path == "/" or path == "/index.html":
            self.serve_html()
        elif path == "/api/status":
            try:
                token = self.get_token()
                valid = False
                if token:
                    valid = self.verify_token(token)
                self.send_json({"logged_in": valid})
            except Exception as e:
                self.send_json({"logged_in": False, "error": str(e)})
        elif path == "/api/auth/me":
            self.auth_me()
        elif path == "/api/auth/modules":
            self.auth_modules()
        elif path == "/api/auth/workspaces":
            self.auth_workspaces_list()
        elif path == "/api/auth/users":
            self.auth_users_list()
        elif path == "/api/hotels":
            user = self._session_user()
            scope = self._campus_scope(user)
            hotels = {k: v for k, v in HOTEL_IDS.items() if not scope or k in scope}
            self.send_json({"hotels": hotels})
        elif path == "/api/rooms/query":
            self.query_rooms(parsed_path.query)
        elif path == "/api/finance/query":
            self.query_finance(parsed_path.query)
        elif path == "/api/occupancy/monthly":
            self.query_monthly_occupancy(parsed_path.query)
        elif path == "/api/occupancy/daily":
            self.query_daily_occupancy(parsed_path.query)
        elif path == "/api/linen/add":
            self.linen_add()
        elif path == "/api/linen/records":
            self.linen_records(parsed_path.query)
        elif path == "/api/linen/inventory":
            self.linen_inventory(parsed_path.query)
        elif path == "/api/linen/summary":
            self.linen_summary(parsed_path.query)
        elif path == "/api/linen/init_inventory":
            self.linen_init_inventory()
        elif path == "/api/occupancy/yesterday":
            self.query_yesterday_occupancy(parsed_path.query)
        elif path == "/api/occupancy/refresh":
            self.occupancy_refresh(parsed_path.query)
        elif path == "/api/residents/snapshot":
            self.residents_snapshot(parsed_path.query)
        elif path == "/api/income/summary":
            self.income_summary(parsed_path.query)
        elif path == "/api/weekly/stats":
            self.weekly_stats(parsed_path.query)
        elif path == "/api/income/records":
            self.income_records(parsed_path.query)
        elif path == "/api/board/compare":
            self.board_compare(parsed_path.query)
        elif path == "/api/board/monthly":
            self.board_monthly(parsed_path.query)
        elif path == "/api/board/import":
            self.board_import()
        elif path == "/api/board/other_rows":
            self.board_other_rows(parsed_path.query)
        elif path == "/api/board/export":
            self.board_export(parsed_path.query)
        elif path == "/api/test":
            self.api_test()
        elif path == "/api/wake/records":
            self.wake_records(parsed_path.query)
        elif path == "/api/wake/save":
            self.wake_save()
        elif path == "/api/battle/zones":
            self.battle_zones(parsed_path.query)
        elif path == "/api/battle/save":
            self.battle_zones_save()
        elif path == "/api/smokefree/rooms":
            self.smoke_free_rooms(parsed_path.query)
        elif path == "/api/roomtypes/rooms":
            self.room_type_rooms(parsed_path.query)
        elif path.startswith("/api/"):
            api_path = path[4:]
            if parsed_path.query:
                api_path += "?" + parsed_path.query
            self.proxy_api(api_path)
        elif path == "/qr-login":
            self.qr_login()
        elif path == "/qr-image":
            self.serve_qr_image()
        elif path == "/qr-status":
            self.qr_status()
        elif path.startswith("/qr-auth/"):
            self.qr_auth(path[9:])
        elif path.endswith((".css", ".js", ".png", ".jpg", ".gif", ".ico", ".svg", ".woff", ".woff2", ".ttf", ".html", ".json")):
            self.serve_static(path)
        else:
            self.send_error(404)

    def do_POST(self):
        from urllib.parse import urlparse
        parsed_path = urlparse(self.path)
        path = parsed_path.path

        # 模块权限校验（登录接口公开）
        if path.startswith("/api/") and path not in ("/api/auth/login", "/api/login"):
            code, msg = self._check_api_access(path)
            if code:
                return self.send_json({"code": code, "msg": msg}, code)

        if path == "/api/auth/login":
            self.auth_login()
        elif path == "/api/auth/me/update":
            self.auth_me_update()
        elif path == "/api/auth/logout":
            self.auth_logout()
        elif path == "/api/auth/workspaces":
            self.auth_workspaces_create()
        elif path == "/api/auth/users":
            self.auth_users_create()
        elif path == "/api/login":
            content_length = int(self.headers.get('Content-Length', 0))
            post_data = self.rfile.read(content_length)
            try:
                data = json.loads(post_data.decode())
                self.handle_login(data)
            except Exception as e:
                self.send_json({"code": 400, "msg": str(e)}, 400)
        elif path == "/api/payroll/parse":
            self.handle_payroll_parse()
        elif path == "/api/payroll/export":
            self.handle_payroll_export()
        elif path == "/api/linen/add":
            self.linen_add()
        elif path == "/api/linen/init_inventory":
            self.linen_init_inventory()
        elif path == "/api/board/import":
            self.board_import()
        elif path == "/api/room/clean":
            self.room_clean()
        elif path == "/api/wake/save":
            self.wake_save()
        elif path == "/api/battle/save":
            self.battle_zones_save()
        elif path == "/api/smokefree/save":
            self.smoke_free_save()
        elif path == "/api/roomtypes/save":
            self.room_type_save()
        elif path.startswith("/api/"):
            api_path = path[4:]
            if parsed_path.query:
                api_path += "?" + parsed_path.query
            self.proxy_api_post(api_path)
        else:
            self.send_error(404)

    def do_PUT(self):
        """支持PUT请求（如打扫房间 breamRoom）"""
        from urllib.parse import urlparse
        parsed_path = urlparse(self.path)
        path = parsed_path.path
        if path.startswith("/api/auth/workspaces/"):
            self.auth_workspaces_update(path[len("/api/auth/workspaces/"):])
        elif path.startswith("/api/auth/users/"):
            self.auth_users_update(path[len("/api/auth/users/"):])
        elif path.startswith("/api/"):
            code, msg = self._check_api_access(path)
            if code:
                return self.send_json({"code": code, "msg": msg}, code)
            api_path = path[4:]
            if parsed_path.query:
                api_path += "?" + parsed_path.query
            self.proxy_api_put(api_path)
        else:
            self.send_error(404)

    def do_DELETE(self):
        """支持DELETE请求（账号/工作区删除，仅管理员）"""
        from urllib.parse import urlparse
        path = urlparse(self.path).path
        if path.startswith("/api/auth/workspaces/"):
            self.auth_workspaces_delete(path[len("/api/auth/workspaces/"):])
        elif path.startswith("/api/auth/users/"):
            self.auth_users_delete(path[len("/api/auth/users/"):])
        else:
            self.send_error(404)

    def proxy_api_post(self, api_path):
        """转发POST请求到AMS（带body）"""
        import urllib.request as urlreq
        token = self.get_token()
        if not token:
            return self.send_json({"code": 401, "msg": "未登录，请先扫码登录"}, 401)

        from urllib.parse import urlparse
        parsed = urlparse(api_path)
        url = f"{AMS_BASE}{parsed.path}"
        if parsed.query:
            url += "?" + parsed.query

        # 读取body
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length) if content_length else b''

        req = urlreq.Request(url, data=body or None, method="POST", headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": self.headers.get('Content-Type', 'application/json'),
            "User-Agent": "Mozilla/5.0"
        })
        try:
            with urlreq.urlopen(req, timeout=30) as resp:
                d = json.loads(resp.read().decode())
            return self.send_json(d)
        except urllib.error.HTTPError as e:
            body_resp = e.read().decode(errors='replace')
            return self.send_json({"code": e.code, "msg": f"AMS错误: {body_resp[:200]}"}, 200)
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"转发失败: {e}"}, 500)

    def proxy_api_put(self, api_path):
        """转发PUT请求到AMS"""
        token = self.get_token()
        if not token:
            return self.send_json({"code": 401, "msg": "未登录，请先扫码登录"}, 401)

        from urllib.parse import urlparse
        parsed = urlparse(api_path)
        url = f"{AMS_BASE}{parsed.path}"
        if parsed.query:
            url += "?" + parsed.query

        # 读取body
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length) if content_length else None

        req = urllib.request.Request(url, data=body, method='PUT', headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json"
        })
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                self.send_json(json.loads(resp.read().decode()))
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            try:
                error_json = json.loads(error_body)
                self.send_json(error_json, e.code)
            except:
                self.send_json({"code": e.code, "msg": error_body}, e.code)
        except Exception as e:
            self.send_json({"code": 500, "msg": str(e)}, 500)

    def send_cors(self, code):
        self.send_response(code)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Authorization, Content-Type")

    def send_json(self, data, code=200):
        self.send_cors(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.end_headers()
        self.wfile.write(json.dumps(data, ensure_ascii=False).encode())

    def query_rooms(self, query_string):
        """查询房间数据"""
        import sqlite3
        import os
        from urllib.parse import parse_qs
        
        db_path = os.path.join(os.path.dirname(__file__), "finance.db")
        if not os.path.exists(db_path):
            self.send_json({"error": "数据库不存在"})
            return
        
        params = parse_qs(query_string)
        hotel_id = params.get("hotel_id", [None])[0]
        floor_name = params.get("floor", [None])[0]
        
        try:
            db = sqlite3.connect(db_path)
            db.row_factory = sqlite3.Row
            c = db.cursor()
            
            query = "SELECT * FROM rooms WHERE 1=1"
            count_query = "SELECT COUNT(*) FROM rooms WHERE 1=1"
            
            params_list = []
            if hotel_id:
                query += " AND hotel_id = ?"
                count_query += " AND hotel_id = ?"
                params_list.append(hotel_id)
            if floor_name:
                query += " AND floor_name LIKE ?"
                count_query += " AND floor_name LIKE ?"
                params_list.append(f"%{floor_name}%")
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                query += scope_sql
                count_query += scope_sql
                params_list += scope_args
            
            query += " ORDER BY hotel_name, floor_name, room_name"
            
            c.execute(query, params_list)
            rooms = [dict(row) for row in c.fetchall()]
            
            c.execute(count_query, params_list)
            total_count = c.fetchone()[0]
            
            # 统计各校区
            stats = {}
            for room in rooms:
                hotel = room["hotel_name"]
                if hotel not in stats:
                    stats[hotel] = {"rooms": 0, "beds": 0}
                stats[hotel]["rooms"] += 1
                stats[hotel]["beds"] += room["bed_count"]
            
            db.close()
            
            self.send_json({
                "rooms": rooms,
                "total_count": total_count,
                "stats": stats
            })
        except Exception as e:
            self.send_json({"error": str(e)})
    # ===== 在住统计（统一口径，yesterday/monthly/daily共用） =====
    OCC_WHERE = ("business_type_name != '' AND business_type_name NOT LIKE '%消费%' "
                 "AND business_type_name != '起步房价' AND business_type_name != '超时房价'")

    def _occ_query(self, select_cols, group_cols, like_prefix, args):
        """统一的在住统计查询
        select_cols: SQL选择列（如 "hotel_id, SUBSTR(create_time,9,2) AS day, COUNT(*) AS cnt"）
        group_cols: GROUP BY列
        like_prefix: create_time LIKE 前缀（日期或月份）
        args: 查询参数
        """
        import sqlite3
        db = sqlite3.connect(os.path.join(os.path.dirname(os.path.abspath(__file__)), 'finance.db'))
        db.row_factory = sqlite3.Row
        c = db.cursor()
        sql = f"SELECT {select_cols} FROM income_flow WHERE create_time LIKE ? AND {self.OCC_WHERE} GROUP BY {group_cols}"
        c.execute(sql, [like_prefix] + args)
        rows = [dict(r) for r in c.fetchall()]
        db.close()
        return rows

    def query_monthly_occupancy(self, query_string):
        """查询本年每月各校区在住人次（从income_flow_monthly预聚合表，毫秒级）
        口径：在住记录（非消费/非起步房价/非超时房价）按 create_time 月份归类，每条=1人次
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            year = params.get('year', ['2026'])[0]
            hotel_id = params.get('hotel_id', [None])[0]

            # 确保预聚合表最新（全量重建，与每日明细同口径）
            self._refresh_monthly_agg()

            conn = self._db()
            c = conn.cursor()
            sql = (f"SELECT hotel_name, ym, cnt FROM income_flow_monthly "
                   f"WHERE ym LIKE ?")
            args = [year + '%']
            if hotel_id:
                sql += " AND hotel_id=?"
                args.append(hotel_id)
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                sql += scope_sql
                args += scope_args
            sql += " ORDER BY ym, hotel_id"
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "records": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    _monthly_agg_ts = 0.0  # 类变量：上次月度聚合重建时间戳，避免每次请求都全量重建

    def _refresh_monthly_agg(self):
        """全量重建月度预聚合表，与「每日在住人数明细」同一口径：
        每天优先取实际快照(daily_residents)，没有快照的日期用收入流水人次补齐，
        再按月汇总，保证两个子模块数据一致。
        5 分钟内重复请求不重建，避免每次打开页面都重算百万行流水。
        """
        import time
        if time.time() - self.__class__._monthly_agg_ts < 300:
            return
        try:
            conn = self._db()
            c = conn.cursor()
            conn.execute("BEGIN IMMEDIATE")
            c.execute("DELETE FROM income_flow_monthly")
            c.execute("""INSERT OR REPLACE INTO income_flow_monthly (ym, hotel_id, hotel_name, cnt, income)
                         WITH income_daily AS (
                             SELECT hotel_id, hotel_name,
                                    SUBSTR(create_time,1,10) AS d,
                                    COUNT(*) AS cnt,
                                    COALESCE(SUM(real_money),0) AS money
                             FROM income_flow
                             WHERE business_type_name != ''
                               AND business_type_name NOT LIKE '%消费%'
                               AND business_type_name != '起步房价'
                               AND business_type_name != '超时房价'
                             GROUP BY hotel_id, hotel_name, d
                         ),
                         merged AS (
                             SELECT i.hotel_id, i.hotel_name, i.d, i.cnt, i.money
                             FROM income_daily i
                             LEFT JOIN daily_residents s
                                    ON s.hotel_id = i.hotel_id AND s.snap_date = i.d
                             WHERE s.snap_date IS NULL
                             UNION ALL
                             SELECT s.hotel_id, s.hotel_name, s.snap_date,
                                    s.resident_count, 0
                             FROM daily_residents s
                         )
                         SELECT SUBSTR(d,1,7) AS ym, hotel_id, hotel_name,
                                SUM(cnt), SUM(money)
                         FROM merged
                         GROUP BY ym, hotel_id, hotel_name""")
            conn.commit()
            conn.close()
            self.__class__._monthly_agg_ts = time.time()
        except Exception as e:
            print(f"[ams_proxy] 月度聚合重建失败: {e}", flush=True)

    def query_daily_occupancy(self, query_string):
        """查询某月每天各校区在住人数。
        优先使用daily_residents实际快照；没有快照的历史日期再用收入流水人次补齐。
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            month = params.get('month', ['2026-08'])[0]  # YYYY-MM
            conn = self._db()
            c = conn.cursor()
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            # 旧日期没有快照时，保留原有收入流水统计作为回退值。
            c.execute(f"""SELECT hotel_id, SUBSTR(create_time, 9, 2) AS day, COUNT(*) AS cnt
                         FROM income_flow WHERE create_time LIKE ? AND business_type_name != ''
                           AND business_type_name NOT LIKE '%消费%' AND business_type_name != '起步房价'
                           AND business_type_name != '超时房价'
                         {scope_sql}
                         GROUP BY hotel_id, day""", [month + '%'] + scope_args)
            merged = {(str(r['hotel_id']), str(r['day']).zfill(2)): int(r['cnt'] or 0) for r in c.fetchall()}
            # 快照是实际在住人数，覆盖同日期同校区的流水回退值。
            c.execute(f"""SELECT hotel_id, SUBSTR(snap_date, 9, 2) AS day, resident_count
                         FROM daily_residents WHERE snap_date LIKE ?{scope_sql}""", [month + '%'] + scope_args)
            snapshot_days = 0
            for r in c.fetchall():
                merged[(str(r['hotel_id']), str(r['day']).zfill(2))] = int(r['resident_count'] or 0)
                snapshot_days += 1
            conn.close()
            rows = [{"hotel_id": hid, "day": day, "cnt": cnt} for (hid, day), cnt in sorted(merged.items())]
            return self.send_json({"code": 200, "records": rows, "source": "daily_snapshot_with_income_fallback", "snapshot_rows": snapshot_days})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def occupancy_refresh(self, query_string):
        """实时刷新在住统计：
        1. 增量拉取最新收入流水（fetch_income_flow_range.py，从最后一条到今天）
        2. 可选 live=hotel_id:count,... 更新今日快照人数为实时值
        3. 强制重建月度预聚合表，返回最新 monthly + daily
        """
        from urllib.parse import parse_qs
        import subprocess, sys
        from datetime import date, timedelta
        params = parse_qs(query_string)
        year = params.get('year', [str(date.today().year)])[0]
        month = params.get('month', [date.today().strftime('%Y-%m')])[0]
        try:
            base = os.path.dirname(os.path.abspath(__file__))
            script = os.path.join(base, 'fetch_income_flow_range.py')
            conn = self._db()
            c = conn.cursor()
            c.execute("SELECT MAX(create_time) FROM income_flow")
            row = c.fetchone()
            conn.close()
            last = (row[0] or '').split(' ')[0]
            if not last:
                last = (date.today() - timedelta(days=1)).strftime('%Y-%m-%d')
            end = date.today().strftime('%Y-%m-%d')
            fetch_log = '流水已是最新'
            if last < end:
                p = subprocess.run([sys.executable, script, last, end],
                                   capture_output=True, text=True, encoding='utf-8', errors='replace',
                                   timeout=900, cwd=base)
                out = (p.stdout or '')
                if p.returncode != 0:
                    out += '\n' + (p.stderr or '')
                    return self.send_json({"code": 500, "msg": "收入流水拉取失败（可能TOKEN过期）", "log": out[-1200:]}, 500)
                fetch_log = out[-800:]

            # 可选：更新今日快照人数为实时值
            live = params.get('live', [None])[0]
            if live:
                today = date.today().strftime('%Y-%m-%d')
                conn = self._db()
                c = conn.cursor()
                for pair in live.split(','):
                    if ':' not in pair:
                        continue
                    hid, cnt = pair.split(':', 1)
                    try:
                        cnt = int(cnt)
                    except ValueError:
                        continue
                    c.execute("UPDATE daily_residents SET resident_count=? WHERE snap_date=? AND hotel_id=?", (cnt, today, hid))
                conn.commit()
                conn.close()

            # 强制重建月度聚合（绕过5分钟缓存）
            self.__class__._monthly_agg_ts = 0.0
            self._refresh_monthly_agg()

            # 查询最新 monthly
            conn = self._db()
            c = conn.cursor()
            c.execute("SELECT hotel_name, ym, cnt FROM income_flow_monthly WHERE ym LIKE ? ORDER BY ym, hotel_id", (year + '%',))
            monthly = [dict(r) for r in c.fetchall()]
            conn.close()

            # 查询最新 daily
            conn = self._db()
            c = conn.cursor()
            c.execute("""SELECT hotel_id, SUBSTR(create_time, 9, 2) AS day, COUNT(*) AS cnt
                         FROM income_flow WHERE create_time LIKE ? AND business_type_name != ''
                           AND business_type_name NOT LIKE '%消费%' AND business_type_name != '起步房价'
                           AND business_type_name != '超时房价'
                         GROUP BY hotel_id, day""", (month + '%',))
            merged = {(str(r['hotel_id']), str(r['day']).zfill(2)): int(r['cnt'] or 0) for r in c.fetchall()}
            c.execute("""SELECT hotel_id, SUBSTR(snap_date, 9, 2) AS day, resident_count
                         FROM daily_residents WHERE snap_date LIKE ?""", (month + '%',))
            for r in c.fetchall():
                merged[(str(r['hotel_id']), str(r['day']).zfill(2))] = int(r['resident_count'] or 0)
            conn.close()
            daily = [{"hotel_id": hid, "day": day, "cnt": cnt} for (hid, day), cnt in sorted(merged.items())]

            return self.send_json({"code": 200, "monthly": monthly, "daily": daily, "log": fetch_log})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    # ===== 布草管理 =====
    def _db(self):
        import os, sqlite3
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'finance.db')
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        return conn

    def linen_add(self):
        """批量录入布草记录（一校区×多物品×4操作）
        body: {date, campus, items: {物品: {运走: n, 运回: n, 破损: n, 返工: n}}, remark}
        校验：运走数不能超过当前库存（自动提示含返工数）
        """
        import json
        try:
            cl = int(self.headers.get('Content-Length', 0))
            data = json.loads(self.rfile.read(cl).decode())
            date = data.get("date")
            campus = data.get("campus")
            items = data.get("items") or {}
            if not date or not campus:
                return self.send_json({"code": 500, "msg": "日期和校区必填"}, 500)

            conn = self._db()
            c = conn.cursor()
            warnings = []
            inserted = 0

            # 计算当前库存（初始 + 运回 - 运走 - 破损）
            def current_stock(campus, item):
                c.execute("SELECT initial_qty FROM linen_inventory WHERE campus=? AND item=?", (campus, item))
                row = c.fetchone()
                initial = row["initial_qty"] if row else 0
                c.execute("""SELECT
                    COALESCE(SUM(CASE WHEN op_type='运回' THEN quantity ELSE 0 END),0) AS back,
                    COALESCE(SUM(CASE WHEN op_type='运走' THEN quantity ELSE 0 END),0) AS out,
                    COALESCE(SUM(CASE WHEN op_type='破损' THEN quantity ELSE 0 END),0) AS broken
                    FROM linen_records WHERE campus=? AND item=?""", (campus, item))
                r = c.fetchone()
                return initial + r["back"] - r["out"] - r["broken"]

            for item, ops in items.items():
                for op, qty in ops.items():
                    qty = int(qty or 0)
                    if qty <= 0:
                        continue
                    # 运走校验：不能超库存；返工自动计入运走
                    if op == "运走":
                        stock = current_stock(campus, item)
                        # 返工累计（该校区该物品所有返工 - 已计入运走的）
                        c.execute("SELECT COALESCE(SUM(quantity),0) FROM linen_records WHERE campus=? AND item=? AND op_type='返工'", (campus, item))
                        rework_total = c.fetchone()[0]
                        if qty + rework_total > stock:
                            warnings.append(f"{item}: 运走{qty}+返工累计{rework_total} = {qty+rework_total} > 库存{stock}，已拦截")
                            continue
                    c.execute("INSERT INTO linen_records (record_date, campus, item, op_type, quantity) VALUES (?,?,?,?,?)",
                              (date, campus, item, op, qty))
                    inserted += 1
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "inserted": inserted, "warnings": warnings})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def linen_records(self, query_string):
        """查询台账（按校区/日期/物品筛选）"""
        import os
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            campus = params.get('campus', [None])[0]
            start = params.get('start', [None])[0]
            end = params.get('end', [None])[0]
            item = params.get('item', [None])[0]

            conn = self._db()
            c = conn.cursor()
            sql = "SELECT * FROM linen_records WHERE 1=1"
            args = []
            if campus:
                sql += " AND campus=?"
                args.append(campus)
            if start:
                sql += " AND record_date>=?"
                args.append(start)
            if end:
                sql += " AND record_date<=?"
                args.append(end)
            if item:
                sql += " AND item=?"
                args.append(item)
            scope_names = self._campus_scope_names(self._session_user())
            if scope_names:
                ph = ",".join("?" * len(scope_names))
                sql += f" AND campus IN ({ph})"
                args += scope_names
            sql += " ORDER BY record_date DESC, id DESC LIMIT 500"
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "records": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def linen_inventory(self, query_string):
        """查询各校区各物品实时库存（初始+运回-运走-破损）"""
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            campus = params.get('campus', [None])[0]

            conn = self._db()
            c = conn.cursor()
            sql = """SELECT inv.campus, inv.item, inv.initial_qty,
                     COALESCE(SUM(CASE WHEN r.op_type='运回' THEN r.quantity ELSE 0 END),0) AS back_qty,
                     COALESCE(SUM(CASE WHEN r.op_type='运走' THEN r.quantity ELSE 0 END),0) AS out_qty,
                     COALESCE(SUM(CASE WHEN r.op_type='破损' THEN r.quantity ELSE 0 END),0) AS broken_qty
                     FROM linen_inventory inv
                     LEFT JOIN linen_records r ON r.campus=inv.campus AND r.item=inv.item"""
            args = []
            where_parts = []
            if campus:
                where_parts.append("inv.campus=?")
                args.append(campus)
            scope_names = self._campus_scope_names(self._session_user())
            if scope_names:
                ph = ",".join("?" * len(scope_names))
                where_parts.append(f"inv.campus IN ({ph})")
                args += scope_names
            if where_parts:
                sql += " WHERE " + " AND ".join(where_parts)
            sql += " GROUP BY inv.campus, inv.item ORDER BY inv.campus, inv.item"
            c.execute(sql, args)
            rows = []
            for r in c.fetchall():
                d = dict(r)
                d["stock"] = d["initial_qty"] + d["back_qty"] - d["out_qty"] - d["broken_qty"]
                rows.append(d)
            conn.close()
            return self.send_json({"code": 200, "records": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def linen_summary(self, query_string):
        """月度对账汇总：各校区×物品×操作合计 + 破损赔偿金额"""
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            month = params.get('month', [None])[0]  # YYYY-MM
            campus = params.get('campus', [None])[0]

            conn = self._db()
            c = conn.cursor()
            sql = """SELECT campus, item, op_type, SUM(quantity) AS total
                     FROM linen_records WHERE 1=1"""
            args = []
            if month:
                sql += " AND record_date LIKE ?"
                args.append(month + '%')
            if campus:
                sql += " AND campus=?"
                args.append(campus)
            scope_names = self._campus_scope_names(self._session_user())
            if scope_names:
                ph = ",".join("?" * len(scope_names))
                sql += f" AND campus IN ({ph})"
                args += scope_names
            sql += " GROUP BY campus, item, op_type"
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]

            # 单价
            c.execute("SELECT item, price FROM linen_prices")
            prices = {r["item"]: r["price"] for r in c.fetchall()}
            # 破损赔偿金额
            for r in rows:
                if r["op_type"] == "破损":
                    r["amount"] = round(r["total"] * prices.get(r["item"], 0), 2)
            conn.close()
            return self.send_json({"code": 200, "records": rows, "prices": prices})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def linen_init_inventory(self):
        """统一录入初始库存（body: {campus: {item: qty}}）"""
        import json
        try:
            cl = int(self.headers.get('Content-Length', 0))
            data = json.loads(self.rfile.read(cl).decode())
            conn = self._db()
            c = conn.cursor()
            for campus, items in data.items():
                for item, qty in items.items():
                    c.execute("INSERT OR REPLACE INTO linen_inventory (campus, item, initial_qty) VALUES (?,?,?)",
                              (campus, item, int(qty)))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "初始库存已保存"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def query_yesterday_occupancy(self, query_string):
        """查询各校区昨日在住人数（从income_flow缓存表统计昨天的在住记录）"""
        from urllib.parse import parse_qs
        import datetime
        try:
            params = parse_qs(query_string)
            date = params.get('date', [None])[0]  # YYYY-MM-DD，默认昨天
            if not date:
                date = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()

            conn = self._db()
            c = conn.cursor()
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            c.execute(f"""SELECT hotel_id, hotel_name, COUNT(*) AS cnt
                         FROM income_flow
                         WHERE create_time LIKE ? AND {self.OCC_WHERE}{scope_sql}
                         GROUP BY hotel_id, hotel_name""", [date + '%'] + scope_args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "date": date, "records": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def residents_snapshot(self, query_string):
        """读取每日在住快照（从daily_residents表），返回指定日期各校区在住人数
        query: date=YYYY-MM-DD（默认最近一天），detail=1 返回完整在住名单（本地缓存，不实时调AMS）
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            date = params.get('date', [None])[0]
            detail = (params.get('detail', ['0'])[0] == '1')

            conn = self._db()
            c = conn.cursor()
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if date:
                c.execute(f"SELECT snap_date, hotel_id, hotel_name, resident_count, raw_json FROM daily_residents WHERE snap_date=?{scope_sql} ORDER BY hotel_id", [date] + scope_args)
            else:
                c.execute(f"""SELECT snap_date, hotel_id, hotel_name, resident_count, raw_json FROM daily_residents
                             WHERE snap_date = (SELECT MAX(snap_date) FROM daily_residents)
                             {scope_sql}
                             ORDER BY hotel_id""", scope_args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()

            if not detail:
                records = [{"snap_date": r["snap_date"], "hotel_id": r["hotel_id"],
                            "hotel_name": r["hotel_name"], "resident_count": r["resident_count"]} for r in rows]
            else:
                import json as _json
                records = []
                for r in rows:
                    people = []
                    try:
                        raw = r.get("raw_json")
                        if raw:
                            people = _json.loads(raw) if isinstance(raw, str) else raw
                    except Exception:
                        people = []
                    for p in people or []:
                        if isinstance(p, dict):
                            p = dict(p)
                            p["_campusName"] = r["hotel_name"]
                            records.append(p)
            return self.send_json({"code": 200, "records": records,
                                   "snap_date": rows[0]["snap_date"] if rows else None})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)
    def income_summary(self, query_string):
        """本地收入汇总（从income_flow表统计，毫秒级，替代AMS实时接口4.5秒）
        query: hotel_id=10144&start=2026-08-01&end=2026-08-11
        返回: {code, income(金额), count(条数)} 兼容AMS的total字段
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            hotel_id = params.get('hotel_id', [None])[0]
            start = params.get('start', [None])[0]
            end = params.get('end', [None])[0]

            conn = self._db()
            c = conn.cursor()
            sql = "SELECT COALESCE(SUM(real_money),0) AS income, COUNT(*) AS cnt FROM income_flow WHERE 1=1"
            args = []
            if hotel_id:
                sql += " AND hotel_id=?"
                args.append(hotel_id)
            if start:
                sql += " AND create_time>=?"
                args.append(start + " 00:00:00")
            if end:
                sql += " AND create_time<=?"
                args.append(end + " 23:59:59")
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                sql += scope_sql
                args += scope_args
            c.execute(sql, args)
            row = c.fetchone()
            conn.close()
            return self.send_json({"code": 200, "income": row["income"], "cnt": row["cnt"],
                                   "total": row["cnt"], "data": {"total": row["income"]}})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def weekly_stats(self, query_string):
        """周效比本地统计（从income_flow表算，毫秒级，替代AMS分页拉取）
        query: hotel_id=10144&start=2026-08-03&end=2026-08-09
        返回: {code, income, total_count, teacher, free, self_pay, occ7, marks:{标记:计数}}
        口径与周效比一致：在住=非空非消费非起步房价非超时房价；
        教师=标记含老师/教练/研发/高报/财务；其余金额0=包住>0=自费
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            hotel_id = params.get('hotel_id', [None])[0]
            start = params.get('start', [None])[0]
            end = params.get('end', [None])[0]

            conn = self._db()  # _db已设置row_factory=sqlite3.Row
            c = conn.cursor()

            where = "create_time>=? AND create_time<=? AND business_type_name != '' AND business_type_name NOT LIKE '%消费%' AND business_type_name != '起步房价' AND business_type_name != '超时房价'"
            args = [start + " 00:00:00", end + " 23:59:59"]
            if hotel_id:
                where += " AND hotel_id=?"
                args.append(hotel_id)
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                where += scope_sql
                args += scope_args

            # 收入（全部流水金额）
            c.execute(f"SELECT COALESCE(SUM(real_money),0) AS income, COUNT(*) AS total FROM income_flow WHERE {where}", args)
            row = c.fetchone()
            income = row["income"] or 0
            total_count = row["total"] or 0

            # 分类统计（教师/包住/自费）
            c.execute(f"""SELECT medi_code_name AS mark, real_money, COUNT(*) AS cnt
                         FROM income_flow WHERE {where} GROUP BY medi_code_name, real_money""", args)
            teacher = free = self_pay = occ7 = 0
            marks = {}
            for r in c.fetchall():
                mark = r["mark"] or ""
                cnt = r["cnt"]
                occ7 += cnt
                if mark:
                    marks[mark] = marks.get(mark, 0) + cnt
                if any(k in mark for k in ['老师', '教练', '研发', '高报', '财务']):
                    teacher += cnt
                elif (r["real_money"] or 0) == 0:
                    free += cnt
                else:
                    self_pay += cnt

            conn.close()
            return self.send_json({"code": 200, "income": income, "total_count": total_count,
                                   "teacher": teacher, "free": free, "self_pay": self_pay,
                                   "occ7": occ7, "marks": marks})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def income_records(self, query_string):
        """本地收入流水明细（从income_flow表，分页，替代AMS实时）
        query: hotel_id=10144&start=2026-08-01&end=2026-08-11&page=1&pageSize=20
        返回: {code, total, rows:[{...}]} 兼容收入报表
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            hotel_id = params.get('hotel_id', [None])[0]
            start = params.get('start', [None])[0]
            end = params.get('end', [None])[0]
            page = int(params.get('page', ['1'])[0])
            page_size = int(params.get('pageSize', ['20'])[0])

            conn = self._db()
            c = conn.cursor()
            where = "1=1"
            args = []
            if hotel_id:
                where += " AND hotel_id=?"
                args.append(hotel_id)
            if start:
                where += " AND create_time>=?"
                args.append(start + " 00:00:00")
            if end:
                where += " AND create_time<=?"
                args.append(end + " 23:59:59")
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                where += scope_sql
                args += scope_args

            # 总数
            c.execute(f"SELECT COUNT(*) FROM income_flow WHERE {where}", args)
            total = c.fetchone()[0]

            # 分页
            offset = (page - 1) * page_size
            c.execute(f"""SELECT hotel_name, room_name, business_type_name, real_money, medi_code_name, order_no, create_time
                         FROM income_flow WHERE {where}
                         ORDER BY create_time DESC LIMIT ? OFFSET ?""", args + [page_size, offset])
            rows = []
            for r in c.fetchall():
                rows.append({
                    "merchantName": r["hotel_name"],
                    "roomName": r["room_name"],
                    "businessTypeName": r["business_type_name"],
                    "realMoney": r["real_money"],
                    "mediCodeName": r["medi_code_name"],
                    "roomOrderNo": r["order_no"],
                    "createTime": r["create_time"],
                })
            conn.close()
            return self.send_json({"code": 200, "total": total, "rows": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def board_compare(self, query_string):
        """包住信息对比查询（从board_compare表）
        query: filter=all|diff|only_ams|only_other&page=1&pageSize=50
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            f = params.get('filter', ['all'])[0]
            page = int(params.get('page', ['1'])[0])
            page_size = int(params.get('pageSize', ['50'])[0])
            exclude_zt = params.get('exclude_zt', ['0'])[0] == '1'
            exclude_zero = params.get('exclude_zero', ['0'])[0] == '1'
            exclude_no_ams = params.get('exclude_no_ams', ['0'])[0] == '1'
            start_month = params.get('start_month', [None])[0]
            end_month = params.get('end_month', [None])[0]

            conn = self._db()
            c = conn.cursor()

            where = "1=1"
            if exclude_zt:
                # 排除直通班：AMS(ams_zt=0) 和 对方(other_zt=0) 都排除
                where = "ams_zt = 0 AND other_zt = 0"
            if exclude_zero:
                # 排除对方天数为0（非直通班的0天记录：面试无限学等）
                where += " AND (other_days > 0 OR other_zt = 1)"
            if exclude_no_ams:
                # 排除"对方有但AMS为0"：AMS=0 且（对方有天数 或 文字天数）
                where += """ AND NOT (ams_days = 0 AND other_exist = 1 AND (other_days > 0 OR (other_zt = 0 AND other_days = 0)))"""
            # 注意：月份筛选只作用于AMS条数卡片(ams_rows)，不影响对比统计
            if f == 'diff':
                where += " AND ams_days != other_days AND ams_days > 0 AND other_days > 0"
            elif f == 'amt_diff':
                where += " AND ams_days = other_days AND other_amt > ams_amt AND ams_days > 0"
            elif f == 'only_ams':
                where += " AND other_days = 0 AND ams_days > 0"
            elif f == 'only_other':
                where += " AND ams_days = 0 AND other_days > 0"
            elif f == 'both':
                where += " AND ams_days > 0 AND other_days > 0"
            elif f == 'text_days':
                # 文字天数：对方有记录(other_exist=1)但天数=0的非直通班（面试无限学等）
                where += " AND other_exist = 1 AND other_zt = 0 AND other_days = 0"

            # 总数
            c.execute(f"SELECT COUNT(*) FROM board_compare WHERE {where}")
            total = c.fetchone()[0]

            # 分页（board_compare已按cert合并：天数加总、校区=主校区）
            offset = (page - 1) * page_size
            c.execute(f"""SELECT bc.cert, bc.campus, bc.name, bc.ams_days, bc.other_days, bc.ams_amt, bc.other_amt,
                         (bc.ams_days - bc.other_days) AS diff_days, bc.other_zt, bc.other_exist,
                         (SELECT GROUP_CONCAT(DISTINCT days_text) FROM board_other_rows WHERE cert=bc.cert AND days_text != '' LIMIT 1) AS days_text
                         FROM board_compare bc WHERE {where}
                         ORDER BY ABS(bc.ams_days - bc.other_days) DESC, bc.name
                         LIMIT ? OFFSET ?""", (page_size, offset))
            page_rows = c.fetchall()
            # 每人分月明细（用于跨月展开，按cert+month聚合）
            c.execute("""SELECT cert, month, campus, SUM(ams_days) ams_days FROM board_compare_monthly
                         GROUP BY cert, month, campus ORDER BY cert, month, campus""")
            person_months = {}
            for r in c.fetchall():
                person_months.setdefault(r['cert'], []).append((r['month'], r['campus'], r['ams_days']))
            rows = []
            for r in page_rows:
                d = dict(r)
                # 分月明细（跨月展开用）
                d['months'] = person_months.get(d['cert'], [])
                rows.append(d)
                # 状态判定：天数+金额综合
                days_same = d['ams_days'] == d['other_days']
                amt_same = abs((d['ams_amt'] or 0) - (d['other_amt'] or 0)) < 0.01
                # 文字天数优先（对方有记录、非直通班、天数=0、有文字标记）
                if d.get('other_exist') == 1 and d.get('other_zt') == 0 and d['other_days'] == 0 and d.get('days_text'):
                    d['status'] = '文字天数'
                elif d['ams_days'] == 0 and d['other_days'] > 0:
                    d['status'] = '仅对方有'
                elif d['other_days'] == 0 and d['ams_days'] > 0:
                    d['status'] = '仅AMS有'
                elif days_same and amt_same:
                    d['status'] = '一致'
                elif days_same and (d['other_amt'] or 0) > (d['ams_amt'] or 0):
                    d['status'] = '金额差异(对方多)'
                else:
                    d['status'] = '不一致'

            # 汇总（按cert去重、分类互斥：直通班>文字天数>有天数分类）
            base_where = where
            c.execute(f"SELECT COUNT(DISTINCT cert) FROM board_compare WHERE {base_where}")
            total_all = c.fetchone()[0]

            # 分类统计（按对方每行cert+收款日期独立判断，一致+不一致+其他=对方条数）
            c.execute(f"""SELECT or2.cert, or2.days, or2.amt, or2.days_text, or2.is_zt,
                         (SELECT MAX(bc.ams_days) FROM board_compare bc WHERE bc.cert=or2.cert) ams_days,
                         (SELECT MAX(bc.ams_amt) FROM board_compare bc WHERE bc.cert=or2.cert) ams_amt
                         FROM board_other_rows or2""")
            other_total = 0
            zt_count = text_days = match = diff = amt_diff = only_other = 0
            for r in c.fetchall():
                # 排除直通班时跳过
                if exclude_zt and r['is_zt'] == 1:
                    continue
                ams_days = r['ams_days'] or 0
                ams_amt = r['ams_amt'] or 0
                other_days = r['days']
                other_amt = r['amt'] or 0
                # 排除"对方有但AMS天数为0"：对方有天数且AMS=0，或文字天数且AMS=0
                if exclude_no_ams and ams_days == 0 and (other_days > 0 or r['days_text']):
                    continue
                other_total += 1
                if r['is_zt'] == 1:
                    zt_count += 1
                elif other_days == 0 and r['days_text']:
                    text_days += 1
                elif ams_days == 0:
                    only_other += 1
                elif ams_days == other_days and abs(ams_amt - other_amt) < 0.01:
                    match += 1
                elif ams_days == other_days and other_amt > ams_amt:
                    amt_diff += 1
                else:
                    diff += 1

            # 对方人数（按cert+收款日期去重：同人不同收款日期算2个）
            zt_filter = "WHERE is_zt=0" if exclude_zt else "WHERE 1=1"
            c.execute(f"""SELECT COUNT(*) FROM (SELECT cert, pay_date FROM board_other_rows
                         {zt_filter} GROUP BY cert, pay_date)""")
            other_person_count = c.fetchone()[0]
            # 对方导入条数（board_other_rows行数，勾选时过滤，且关系）
            # 排除直通班：is_zt=0；排除对方有AMS为0：days>0且cert无AMS
            zt_sql = "AND is_zt=0" if exclude_zt else ""
            if exclude_no_ams:
                # 排除：AMS为0 且（对方有天数 或 文字天数）
                c.execute(f"""SELECT COUNT(*) FROM board_other_rows or2
                             WHERE 1=1 {zt_sql} AND NOT (NOT EXISTS (SELECT 1 FROM board_compare bc 
                                               WHERE bc.cert=or2.cert AND bc.ams_days > 0) AND (or2.days > 0 OR or2.days_text != ''))""")
                other_rows = c.fetchone()[0]
            elif exclude_zt:
                c.execute("SELECT COUNT(*) FROM board_other_rows WHERE is_zt=0")
                other_rows = c.fetchone()[0]
            elif exclude_zero:
                c.execute("SELECT COUNT(*) FROM board_other_rows WHERE days>0 OR is_zt=1")
                other_rows = c.fetchone()[0]
            else:
                c.execute("SELECT COUNT(*) FROM board_other_rows")
                other_rows = c.fetchone()[0]
            # AMS条数（board_compare_monthly，按月份筛选）
            month_where = " WHERE 1=1"
            if start_month or end_month:
                sm = start_month or '01'
                em = end_month or '12'
                month_where = f" WHERE substr(month,6,2) >= '{sm}' AND substr(month,6,2) <= '{em}'"
            if exclude_zt:
                # AMS直通班标记在board_compare，monthly表没有 → 用board_compare反查
                c.execute(f"""SELECT COUNT(*) FROM board_compare_monthly m{month_where}
                             AND NOT EXISTS (SELECT 1 FROM board_compare bc WHERE bc.cert=m.cert AND bc.campus=m.campus AND bc.ams_zt=1)""")
            else:
                c.execute(f"SELECT COUNT(*) FROM board_compare_monthly m{month_where}")
            ams_rows = c.fetchone()[0]

            conn.close()
            return self.send_json({"code": 200, "total": total, "rows": rows,
                                   "summary": {"total_all": total_all, "other_total": other_total,
                                               "other_person_count": other_person_count,
                                               "match": match, "diff": diff,
                                               "amt_diff": amt_diff, "only_other": only_other,
                                               "text_days": text_days, "zt_count": zt_count,
                                               "other_rows": other_rows, "ams_rows": ams_rows}})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def board_monthly(self, query_string):
        """包住分月明细（按身份证查各月AMS天数）
        query: cert=150xxx
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            cert = params.get('cert', [None])[0]
            if not cert:
                return self.send_json({"code": 400, "msg": "缺少cert参数"}, 400)
            conn = self._db()
            c = conn.cursor()
            c.execute("SELECT month, ams_days, campus FROM board_compare_monthly WHERE cert=? ORDER BY month", (cert,))
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "records": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def room_clean(self):
        """打扫房间：调用AMS breamRoom接口（脏房→干净房，会写日志）
        body: {hotelId, roomCode} 或 {hotelId, roomName}
        """
        import urllib.request as urlreq
        content_length = int(self.headers.get('Content-Length', 0))
        try:
            data = json.loads(self.rfile.read(content_length).decode())
        except Exception:
            return self.send_json({"code": 400, "msg": "参数错误"}, 400)

        hotel_id = data.get('hotelId') or data.get('hotel_id')
        room_code = data.get('roomCode') or data.get('room_code')
        room_name = data.get('roomName') or data.get('room_name')

        if not hotel_id:
            return self.send_json({"code": 400, "msg": "缺少hotelId"}, 400)

        token = self.get_token()
        if not token:
            return self.send_json({"code": 401, "msg": "未登录，请先扫码登录"}, 401)

        # 如果给了roomName，先查roomCode
        if not room_code and room_name:
            try:
                conn = self._db()
                c = conn.cursor()
                c.execute("SELECT roomCode FROM room_data WHERE hotelId=? AND roomName LIKE ?", (hotel_id, f"%{room_name}%"))
                r = c.fetchone()
                conn.close()
                if r:
                    room_code = r[0]
                else:
                    return self.send_json({"code": 404, "msg": f"未找到房间: {room_name}"}, 404)
            except Exception:
                pass  # 本地没room_data表，交给AMS按名称查

        if not room_code:
            return self.send_json({"code": 400, "msg": "缺少roomCode或roomName"}, 400)

        # 调AMS breamRoom
        url = f"{AMS_BASE}/hotel/web/basics/room/breamRoom/{hotel_id}"
        payload = json.dumps({"hotelId": hotel_id, "roomCode": room_code}).encode()
        req = urlreq.Request(url, data=payload, method="PUT", headers={
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "User-Agent": "Mozilla/5.0"
        })
        try:
            with urlreq.urlopen(req, timeout=30) as resp:
                d = json.loads(resp.read().decode())
            return self.send_json(d)
        except urllib.error.HTTPError as e:
            body = e.read().decode(errors='replace')
            return self.send_json({"code": e.code, "msg": f"AMS错误: {body[:200]}"}, 200)
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"打扫失败: {e}"}, 500)

    def board_import(self):
        """导入对方Excel（上传文件，解析身份证/姓名/包住天数，更新对比数据）
        请求: multipart/form-data, file字段=Excel文件
        """
        import cgi
        import tempfile
        import os
        try:
            form = cgi.FieldStorage(
                fp=self.rfile,
                headers=self.headers,
                environ={
                    'REQUEST_METHOD': 'POST',
                    'CONTENT_TYPE': self.headers.get('Content-Type', ''),
                })
            file_item = form['file']
            if not file_item.filename:
                return self.send_json({"code": 400, "msg": "未上传文件"}, 400)

            # 保存临时文件
            tmp_path = os.path.join(tempfile.gettempdir(), 'board_import.xlsx')
            with open(tmp_path, 'wb') as f:
                f.write(file_item.file.read())

            # 解析Excel（每行独立存储，不合并）
            import openpyxl
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            ws = wb[wb.sheetnames[0]]

            # 建分行表（含days_text原始文字 + 收款日期）
            conn = self._db()
            c = conn.cursor()
            c.execute("""CREATE TABLE IF NOT EXISTS board_other_rows (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cert TEXT, name TEXT, days INTEGER DEFAULT 0, amt REAL DEFAULT 0,
                days_text TEXT DEFAULT '',
                is_zt INTEGER DEFAULT 0,
                pay_date TEXT DEFAULT '',
                created_at TEXT DEFAULT (datetime('now','localtime'))
            )""")
            # 兼容旧表（加days_text/is_zt/pay_date列）
            cols = [r[1] for r in c.execute("PRAGMA table_info(board_other_rows)")]
            if 'days_text' not in cols:
                c.execute("ALTER TABLE board_other_rows ADD COLUMN days_text TEXT DEFAULT ''")
            if 'is_zt' not in cols:
                c.execute("ALTER TABLE board_other_rows ADD COLUMN is_zt INTEGER DEFAULT 0")
            if 'pay_date' not in cols:
                c.execute("ALTER TABLE board_other_rows ADD COLUMN pay_date TEXT DEFAULT ''")
            # 清空旧数据
            c.execute("DELETE FROM board_other_rows")
            c.execute("UPDATE board_compare SET other_days=0, other_amt=0, other_zt=0, other_exist=0")

            row_count = 0
            other_total = {}
            other_zt = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                cert = str(row[2]) if row[2] else ''   # 身份证（第3列）
                name = row[3] if row[3] else ''         # 姓名（第4列）
                days_raw = row[8]                       # 包住天数（第9列）
                amt = row[7] or 0                       # 代收住宿费（第8列）
                pay_date = str(row[14])[:10] if row[14] else ''  # 收款日期（第15列）
                # 标记直通班（不排除，前端勾选时过滤）
                is_zt = 1 if (isinstance(days_raw, str) and '直通班' in days_raw) else 0
                days = int(days_raw) if isinstance(days_raw, (int, float)) else 0
                days_text = str(days_raw) if isinstance(days_raw, str) else ''
                if cert:
                    # 分行存储（不合并）
                    c.execute("INSERT INTO board_other_rows (cert, name, days, amt, days_text, is_zt, pay_date) VALUES (?,?,?,?,?,?,?)",
                              (cert, name, days, amt if isinstance(amt, (int, float)) else 0, days_text, is_zt, pay_date))
                    row_count += 1
                    # 累计（对比用）
                    if cert not in other_total:
                        other_total[cert] = {"name": name, "days": 0, "amt": 0}
                    other_total[cert]["days"] += days
                    other_total[cert]["amt"] += amt if isinstance(amt, (int, float)) else 0
                    if is_zt:
                        other_zt[cert] = 1

            os.remove(tmp_path)

            # 更新累计表的对方数据
            for cert, info in other_total.items():
                zt = other_zt.get(cert, 0)
                c.execute("SELECT 1 FROM board_compare WHERE cert=?", (cert,))
                if c.fetchone():
                    c.execute("UPDATE board_compare SET other_days=?, other_amt=?, other_zt=?, other_exist=1 WHERE cert=?", (info["days"], info["amt"], zt, cert))
                else:
                    c.execute("INSERT INTO board_compare (cert, campus, name, ams_days, other_days, ams_amt, other_amt, other_zt, other_exist) VALUES (?,?,?,0,?,0,?,?,1)",
                              (cert, '', info["name"], info["days"], info["amt"], zt))
            conn.commit()
            conn.close()

            return self.send_json({"code": 200, "msg": f"导入成功: {row_count}行, {len(other_total)}人"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"导入失败: {str(e)}"}, 500)

    def api_test(self):
        """测试模块：返回系统状态（版本/时间/数据库）"""
        import datetime
        import sqlite3
        try:
            db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'finance.db')
            db_exists = os.path.exists(db_path)
            db_size = os.path.getsize(db_path) if db_exists else 0
            tables = {}
            if db_exists:
                conn = sqlite3.connect(db_path)
                c = conn.cursor()
                c.execute("SELECT name FROM sqlite_master WHERE type='table'")
                for r in c.fetchall():
                    try:
                        c2 = conn.cursor()
                        c2.execute(f"SELECT COUNT(*) FROM {r[0]}")
                        tables[r[0]] = c2.fetchone()[0]
                    except:
                        pass
                conn.close()
            self.send_json({
                "code": 200,
                "module": "测试模块",
                "version": "v1.0.0",
                "time": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "server": "workbench协作版",
                "database": {"exists": db_exists, "size_mb": round(db_size/1024/1024, 1), "tables": tables}
            })
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)
    def board_other_rows(self, query_string):
        """对方表分行明细（按身份证查所有行）
        query: cert=150xxx
        """
        from urllib.parse import parse_qs
        try:
            params = parse_qs(query_string)
            cert = params.get('cert', [None])[0]
            if not cert:
                return self.send_json({"code": 400, "msg": "缺少cert参数"}, 400)
            conn = self._db()
            c = conn.cursor()
            c.execute("SELECT id, cert, name, days, amt, days_text, is_zt FROM board_other_rows WHERE cert=? ORDER BY id", (cert,))
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "records": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def _wake_db(self):
        """叫醒记录数据库（独立文件，避免和finance.db混）"""
        import sqlite3
        db_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "wake_data.db")
        conn = sqlite3.connect(db_path)
        c = conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS wake_records (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            record_date TEXT NOT NULL,
            hotel_id TEXT,
            room_code TEXT,
            room_name TEXT,
            bed_name TEXT,
            student_name TEXT,
            gender TEXT,
            wake_status TEXT,
            operator TEXT,
            updated_at TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS battle_zones (
            hotel_id TEXT PRIMARY KEY,
            floors TEXT,
            beds TEXT
        )""")
        # 旧库补充 beds 列（床位级战斗区）
        c.execute("""CREATE TABLE IF NOT EXISTS smoke_free (
            hotel_id TEXT PRIMARY KEY,
            rooms TEXT
        )""")
        c.execute("""CREATE TABLE IF NOT EXISTS room_types (
            hotel_id TEXT PRIMARY KEY,
            rooms TEXT
        )""")
        try:
            c.execute("ALTER TABLE battle_zones ADD COLUMN beds TEXT")
            conn.commit()
        except Exception:
            pass
        conn.commit()
        return conn

    def wake_records(self, query_string):
        """查询叫醒记录：?date=2026-08-17&hotel_id=10145 或 ?start_date=&end_date=&hotel_id=10145"""
        from urllib.parse import parse_qs
        params = parse_qs(query_string)
        date = params.get('date', [''])[0]
        hotel_id = params.get('hotel_id', [''])[0]
        start_date = params.get('start_date', [''])[0]
        end_date = params.get('end_date', [''])[0]
        try:
            import sqlite3
            conn = self._wake_db()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            sql = "SELECT * FROM wake_records WHERE 1=1"
            args = []
            if date:
                sql += " AND record_date = ?"
                args.append(date)
            if hotel_id:
                sql += " AND hotel_id = ?"
                args.append(hotel_id)
            if start_date:
                sql += " AND record_date >= ?"
                args.append(start_date)
            if end_date:
                sql += " AND record_date <= ?"
                args.append(end_date)
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                sql += scope_sql
                args += scope_args
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "data": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def wake_save(self):
        """保存叫醒记录（单条upsert：日期+房间+床位唯一）"""
        import json as json_mod
        content_length = int(self.headers.get('Content-Length', 0))
        try:
            d = json_mod.loads(self.rfile.read(content_length).decode())
        except Exception:
            return self.send_json({"code": 400, "msg": "参数错误"}, 400)
        try:
            conn = self._wake_db()
            c = conn.cursor()
            # 删除旧记录（同日期+同房间+同床位）
            c.execute("DELETE FROM wake_records WHERE record_date=? AND room_code=? AND bed_name=?",
                      (d.get('date'), d.get('room_code'), d.get('bed_name')))
            c.execute("""INSERT INTO wake_records (record_date, hotel_id, room_code, room_name, bed_name, student_name, gender, wake_status, operator, updated_at)
                         VALUES (?,?,?,?,?,?,?,?,?,?)""",
                      (d.get('date'), d.get('hotel_id'), d.get('room_code'), d.get('room_name'),
                       d.get('bed_name'), d.get('student_name'), d.get('gender'),
                       d.get('wake_status'), d.get('operator', ''), d.get('updated_at', '')))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "保存成功"})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return self.send_json({"code": 500, "msg": str(e)}, 500)
    def battle_zones(self, query_string):
        """查询战斗区配置：?hotel_id=10145 或全部"""
        from urllib.parse import parse_qs
        params = parse_qs(query_string)
        hotel_id = params.get('hotel_id', [''])[0]
        try:
            import sqlite3
            conn = self._wake_db()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            sql = "SELECT * FROM battle_zones WHERE 1=1"
            args = []
            if hotel_id:
                sql += " AND hotel_id=?"
                args.append(hotel_id)
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                sql += scope_sql
                args += scope_args
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "data": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def battle_zones_save(self):
        """保存战斗区配置：{hotel_id, floors:[5,8], beds:["RFJ..."]}"""
        import json as json_mod
        content_length = int(self.headers.get('Content-Length', 0))
        try:
            d = json_mod.loads(self.rfile.read(content_length).decode())
        except Exception:
            return self.send_json({"code": 400, "msg": "参数错误"}, 400)
        try:
            conn = self._wake_db()
            c = conn.cursor()
            c.execute("DELETE FROM battle_zones WHERE hotel_id=?", (d.get('hotel_id'),))
            c.execute("INSERT INTO battle_zones (hotel_id, floors, beds) VALUES (?,?,?)",
                      (d.get('hotel_id'), json_mod.dumps(d.get('floors', []), ensure_ascii=False),
                       json_mod.dumps(d.get('beds', []), ensure_ascii=False)))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "保存成功"})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return self.send_json({"code": 500, "msg": str(e)}, 500)


    def smoke_free_rooms(self, query_string):
        """查询无烟宿舍配置：?hotel_id=10144 或全部"""
        from urllib.parse import parse_qs
        params = parse_qs(query_string)
        hotel_id = params.get('hotel_id', [''])[0]
        try:
            import sqlite3
            conn = self._wake_db()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            sql = "SELECT * FROM smoke_free WHERE 1=1"
            args = []
            if hotel_id:
                sql += " AND hotel_id=?"
                args.append(hotel_id)
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                sql += scope_sql
                args += scope_args
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "data": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def smoke_free_save(self):
        """保存无烟宿舍配置：{hotel_id, rooms:[房间码...]}"""
        import json as json_mod
        content_length = int(self.headers.get('Content-Length', 0))
        try:
            d = json_mod.loads(self.rfile.read(content_length).decode())
        except Exception:
            return self.send_json({"code": 400, "msg": "参数错误"}, 400)
        try:
            conn = self._wake_db()
            c = conn.cursor()
            c.execute("DELETE FROM smoke_free WHERE hotel_id=?", (d.get('hotel_id'),))
            c.execute("INSERT INTO smoke_free (hotel_id, rooms) VALUES (?,?)",
                      (d.get('hotel_id'), json_mod.dumps(d.get('rooms', []), ensure_ascii=False)))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "保存成功"})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def room_type_rooms(self, query_string):
        """查询房间类型配置：?hotel_id=10144 或全部"""
        from urllib.parse import parse_qs
        params = parse_qs(query_string)
        hotel_id = params.get('hotel_id', [''])[0]
        try:
            import sqlite3
            conn = self._wake_db()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()
            sql = "SELECT * FROM room_types WHERE 1=1"
            args = []
            if hotel_id:
                sql += " AND hotel_id=?"
                args.append(hotel_id)
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "hotel_id")
            if scope_args:
                sql += scope_sql
                args += scope_args
            c.execute(sql, args)
            rows = [dict(r) for r in c.fetchall()]
            conn.close()
            return self.send_json({"code": 200, "data": rows})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def room_type_save(self):
        """保存房间类型配置：{hotel_id, rooms:{房间码:类型}}"""
        import json as json_mod
        content_length = int(self.headers.get('Content-Length', 0))
        try:
            d = json_mod.loads(self.rfile.read(content_length).decode())
        except Exception:
            return self.send_json({"code": 400, "msg": "参数错误"}, 400)
        try:
            conn = self._wake_db()
            c = conn.cursor()
            c.execute("DELETE FROM room_types WHERE hotel_id=?", (d.get('hotel_id'),))
            c.execute("INSERT INTO room_types (hotel_id, rooms) VALUES (?,?)",
                      (d.get('hotel_id'), json_mod.dumps(d.get('rooms', {}), ensure_ascii=False)))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "保存成功"})
        except Exception as e:
            try: conn.close()
            except Exception: pass
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def board_export(self, query_string):
        """导出包住对比（结合AMS格式明细+对比结果）
        Sheet1汇总 / Sheet2对比明细(含身份证手机号房间号) / Sheet3一致 / Sheet4不一致(含原因)
        query: exclude_zt=0|1&exclude_zero=0|1&start_month=&end_month=
        """
        from urllib.parse import parse_qs
        import sqlite3
        try:
            params = parse_qs(query_string)
            exclude_zt = params.get('exclude_zt', ['0'])[0] == '1'
            exclude_zero = params.get('exclude_zero', ['0'])[0] == '1'
            exclude_no_ams = params.get('exclude_no_ams', ['0'])[0] == '1'
            start_month = params.get('start_month', [None])[0]
            end_month = params.get('end_month', [None])[0]

            conn = self._db()
            conn.row_factory = sqlite3.Row
            c = conn.cursor()

            # 过滤条件（与页面一致）
            where = "bc.other_exist=1"
            if exclude_zt:
                where += " AND bc.ams_zt = 0 AND bc.other_zt = 0"
            if exclude_zero:
                where += " AND (bc.other_days > 0 OR bc.other_zt = 1)"
            if exclude_no_ams:
                # 排除"对方有但AMS为0"：AMS=0 且（对方有天数 或 文字天数）
                where += """ AND NOT (bc.ams_days = 0 AND bc.other_exist = 1 AND (bc.other_days > 0 OR (bc.other_zt = 0 AND bc.other_days = 0)))"""

            # 对比明细（按cert+校区分行，同名排一起）——各校区数据从board_compare_monthly汇总
            c.execute(f"""SELECT bc.cert, bc.name, bc.campus, bc.ams_days AS ams_days_total, bc.other_days, bc.ams_amt, bc.other_amt,
                         bc.other_zt, bc.other_exist,
                         (SELECT GROUP_CONCAT(DISTINCT days_text) FROM board_other_rows WHERE cert=bc.cert AND days_text != '' LIMIT 1) AS days_text,
                         (SELECT GROUP_CONCAT(DISTINCT campus) FROM board_compare WHERE cert=bc.cert AND campus != '') AS campuses,
                         d.mobile, d.room_name, d.room_price_code, d.medi_code, d.free_days_list,
                         d.check_in_time, d.due_check_out_time
                         FROM board_compare bc
                         LEFT JOIN (SELECT cert, MAX(mobile) mobile, MAX(room_name) room_name, MAX(room_price_code) room_price_code,
                                    MAX(medi_code) medi_code, MAX(free_days_list) free_days_list,
                                    MAX(check_in_time) check_in_time, MAX(due_check_out_time) due_check_out_time
                                    FROM board_ams_detail GROUP BY cert) d ON d.cert = bc.cert
                         WHERE bc.other_exist=1 AND {where}
                         ORDER BY bc.name""")
            rows = c.fetchall()
            # 各校区明细（同一cert按校区展开）
            campus_rows = {}  # cert -> [(campus, days), ...]
            c.execute("""SELECT cert, campus, SUM(ams_days) days FROM board_compare_monthly
                         GROUP BY cert, campus ORDER BY campus""")
            for r in c.fetchall():
                campus_rows.setdefault(r['cert'], []).append((r['campus'], r['days']))

            # 每人各月明细（cert+月 → 天数/校区）
            cert_months = {}  # cert -> {month: [(campus, days), ...]}
            c.execute("""SELECT cert, month, campus, ams_days FROM board_compare_monthly ORDER BY cert, month, campus""")
            for r in c.fetchall():
                cert_months.setdefault(r['cert'], {}).setdefault(r['month'], []).append((r['campus'], r['ams_days']))

            # 各校区房间号/入住退宿（cert+校区 → 房间信息）
            campus_rooms = {}  # cert -> {campus: (room, price, medi, free_list, cin, cout)}
            c.execute("""SELECT cert, hotel_name, room_name, room_price_code, medi_code, free_days_list,
                         check_in_time, due_check_out_time FROM board_ams_detail
                         ORDER BY cert, hotel_name, check_in_time""")
            for r in c.fetchall():
                cert = r['cert']
                campus = r['hotel_name']
                if cert not in campus_rooms:
                    campus_rooms[cert] = {}
                if campus not in campus_rooms[cert]:
                    campus_rooms[cert][campus] = (r['room_name'], r['room_price_code'], r['medi_code'],
                                                  r['free_days_list'], r['check_in_time'], r['due_check_out_time'])
            conn.close()

            import openpyxl
            from openpyxl.styles import Font, PatternFill
            wb = openpyxl.Workbook()
            # Sheet1 汇总
            ws_sum = wb.active
            ws_sum.title = "汇总"
            ws_sum.append(["分类", "人数"])

            # Sheet2 对比明细（按人分组：单月1行，跨月分月行+合计行）
            MONTHS = ["2026-04", "2026-05", "2026-06", "2026-07", "2026-08"]
            det_headers = ["姓名", "身份证号", "手机号", "校区",
                           "1月", "2月", "3月", "4月", "5月", "6月", "7月", "8月", "9月", "10月", "11月", "12月",
                           "合计", "对方天数", "天数差",
                           "AMS金额", "对方金额", "金额差", "状态", "原因",
                           "最终天数", "最终金额"]
            ws_det = wb.create_sheet("对比明细")
            ws_det.append(det_headers)

            # Sheet 一致 / 不一致（全量）
            ws_m = wb.create_sheet("一致")
            ws_m.append(["姓名", "身份证号", "手机号", "校区", "月份", "AMS天数", "对方天数", "AMS金额", "对方金额"])
            ws_d = wb.create_sheet("不一致")
            ws_d.append(["姓名", "身份证号", "手机号", "校区", "月份", "AMS天数", "对方天数", "天数差",
                         "AMS金额", "对方金额", "金额差", "原因"])

            for ws in (ws_det, ws_m, ws_d):
                for col in range(1, ws.max_column + 1):
                    ws.cell(row=1, column=col).font = Font(bold=True)
                    ws.cell(row=1, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            match_count = 0
            diff_count = 0
            for r in rows:
                d = dict(r)
                # 清理 Excel/AMS 导入姓名中的前导制表符和空格，避免导出后姓名视觉上混在一起。
                export_name = str(d.get('name') or '').strip()
                days_same = d['ams_days_total'] == d['other_days']
                amt_same = abs((d['ams_amt'] or 0) - (d['other_amt'] or 0)) < 0.01
                # 原因判定（用合并后的总天数）
                if d.get('other_zt') == 1:
                    reason = '对方直通班'
                    status = '直通班'
                elif d.get('other_exist') == 1 and d.get('other_days') == 0 and d.get('days_text'):
                    reason = '对方文字天数(' + d['days_text'] + ')'
                    status = '文字天数'
                elif d['ams_days_total'] == 0 and d['other_days'] > 0:
                    reason = 'AMS无此人'
                    status = '仅对方有'
                elif days_same and amt_same:
                    reason = ''
                    status = '一致'
                elif days_same and (d['other_amt'] or 0) > (d['ams_amt'] or 0):
                    reason = '对方金额多 ¥' + str(round((d['other_amt'] or 0) - (d['ams_amt'] or 0)))
                    status = '金额差异(对方多)'
                elif not days_same:
                    reason = '天数不一致(AMS ' + str(d['ams_days_total']) + ' vs 对方 ' + str(d['other_days']) + ')'
                    status = '天数不一致'
                else:
                    reason = 'AMS金额多 ¥' + str(round((d['ams_amt'] or 0) - (d['other_amt'] or 0)))
                    status = '金额差异(AMS多)'

                other_display = d['other_days']
                if d.get('other_zt') == 1:
                    other_display = '直通班'
                elif status == '文字天数' and d.get('days_text'):
                    other_display = d['days_text']

                # 最终天数/金额计算
                # 面试无限学 → 最终=AMS
                # 对方(天数或金额) > AMS → 最终=AMS
                # AMS > 对方 → 留空
                # 相等 → 显示
                is_mianshi = d.get('days_text') and '面试无限学' in str(d.get('days_text'))
                ams_days_t = d['ams_days_total'] or 0
                ams_amt_t = d['ams_amt'] or 0
                other_days_t = d['other_days'] or 0
                other_amt_t = d['other_amt'] or 0
                if is_mianshi:
                    final_days = ams_days_t
                    final_amt = ams_amt_t
                elif other_days_t > ams_days_t or other_amt_t > ams_amt_t:
                    final_days = ams_days_t
                    final_amt = ams_amt_t
                elif ams_days_t > other_days_t or ams_amt_t > other_amt_t:
                    final_days = ''
                    final_amt = ''
                else:
                    final_days = ams_days_t
                    final_amt = ams_amt_t

                # 各月明细（按cert分配；单月1行，跨月分月行+合计行；对方数据只在首行显示）
                person_months = cert_months.get(d['cert'], {})
                if not person_months:
                    person_months = {MONTHS[0]: campus_rows.get(d['cert'], [(d.get('campus') or '', d['ams_days_total'])])}

                # 收集所有月行
                month_rows = []  # (month, campus, days)
                for month in MONTHS:
                    for campus_name, campus_days in person_months.get(month, []):
                        if campus_days > 0:
                            month_rows.append((month, campus_name, campus_days))

                # 每人一行：1-12月分列 + 合计（名字只出现一次）
                month_days_map = {m: 0 for m in range(1, 13)}
                for month, campus_name, campus_days in month_rows:
                    mn = int(month[5:7])
                    if mn in month_days_map:
                        month_days_map[mn] += campus_days
                # 主校区（天数最多的）
                if month_rows:
                    main_campus = max(month_rows, key=lambda x: x[2])[1]
                else:
                    main_campus = d.get('campus') or ''
                campus_show = (main_campus or '').replace('上岸公寓', '').replace('校区', '')
                month_cells = [month_days_map[m] if month_days_map[m] > 0 else '' for m in range(1, 13)]
                ws_det.append([export_name, d.get('cert', ''), d.get('mobile', ''), campus_show,
                               *month_cells,
                               d['ams_days_total'], other_display, (d['ams_days_total'] or 0) - (d['other_days'] or 0),
                               round(d['ams_amt'] or 0), round(d['other_amt'] or 0),
                               round((d['other_amt'] or 0) - (d['ams_amt'] or 0)), status, reason,
                               final_days, final_amt if final_amt == '' else round(final_amt)])

                if status == '一致':
                    match_count += 1
                    ws_m.append([export_name, d.get('cert', ''), d.get('mobile', ''), d.get('campus', '').replace('上岸公寓', '').replace('校区', ''),
                                 d.get('room_name', ''), d['ams_days_total'], d['other_days'], round(d['ams_amt'] or 0), round(d['other_amt'] or 0)])
                else:
                    diff_count += 1
                    ws_d.append([export_name, d.get('cert', ''), d.get('mobile', ''), d.get('campus', '').replace('上岸公寓', '').replace('校区', ''),
                                 d.get('room_name', ''), d['ams_days_total'], d['other_days'], (d['ams_days_total'] or 0) - (d['other_days'] or 0),
                                 round(d['ams_amt'] or 0), round(d['other_amt'] or 0),
                                 round((d['other_amt'] or 0) - (d['ams_amt'] or 0)), reason])

            ws_sum.append(["一致", match_count])
            ws_sum.append(["不一致", diff_count])
            ws_sum.append(["合计", match_count + diff_count])

            for col in range(1, 3):
                ws_sum.cell(row=1, column=col).font = Font(bold=True)
                ws_sum.cell(row=1, column=col).fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

            import io
            buf = io.BytesIO()
            wb.save(buf)
            data = buf.getvalue()
            self.send_response(200)
            self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", 'attachment; filename="board_compare_' + str(match_count) + '_' + str(diff_count) + '.xlsx"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return None
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"导出失败: {str(e)}"}, 500)

    def query_finance(self, query_string):
        """查询财务数据"""
        import sqlite3
        import os
        from urllib.parse import parse_qs
        
        db_path = os.path.join(os.path.dirname(__file__), 'finance.db')
        if not os.path.exists(db_path):
            self.send_json({"error": "数据库不存在"})
            return
        
        params = parse_qs(query_string)
        company = params.get('company', [None])[0]
        data_type = params.get('type', [None])[0]
        start_date = params.get('start', [None])[0]
        end_date = params.get('end', [None])[0]
        
        try:
            db = sqlite3.connect(db_path)
            db.row_factory = sqlite3.Row
            c = db.cursor()
            
            table = data_type if data_type else 'expenses'
            if table not in ['expenses', 'income', 'assets']:
                table = 'expenses'
            
            query = f"SELECT * FROM {table} WHERE 1=1"
            count_query = f"SELECT COUNT(*) FROM {table} WHERE 1=1"
            sum_column = "total_price" if table == "assets" else "amount"
            sum_query = f"SELECT COALESCE(SUM({sum_column}), 0) FROM {table} WHERE 1=1"
            
            params_list = []
            # 校区范围过滤（资产表无campus字段，仅过滤收支表）
            scope_sql, scope_args = self._campus_filter_sql(self._session_user(), "campus")
            if scope_args and table in ("expenses", "income"):
                query += scope_sql
                count_query += scope_sql
                sum_query += scope_sql
                params_list += scope_args
            if company:
                # 前端传的是校区名（B座/C座等），数据库用 campus 字段过滤
                # 同时兼容 company 字段（公司简称），双条件匹配
                query += " AND (campus = ? OR company LIKE ?)"
                count_query += " AND (campus = ? OR company LIKE ?)"
                sum_query += " AND (campus = ? OR company LIKE ?)"
                params_list.append(company)
                params_list.append(f"%{company}%")
            if start_date:
                query += " AND date >= ?"
                count_query += " AND date >= ?"
                sum_query += " AND date >= ?"
                params_list.append(start_date)
            if end_date:
                query += " AND date <= ?"
                count_query += " AND date <= ?"
                sum_query += " AND date <= ?"
                params_list.append(end_date)
            
            query += " ORDER BY date DESC LIMIT 500"
            
            c.execute(query, params_list)
            records = [dict(row) for row in c.fetchall()]
            
            c.execute(count_query, params_list)
            total_count = c.fetchone()[0]
            
            c.execute(sum_query, params_list)
            total_amount = c.fetchone()[0]
            
            db.close()
            
            self.send_json({
                "records": records,
                "total_count": total_count,
                "total_amount": total_amount
            })
        except Exception as e:
            self.send_json({"error": str(e)})

    def serve_html(self):
        try:
            with open(HTML_PATH, "r", encoding="utf-8") as f:
                html = f.read()
            self.send_cors(200)
            self.send_header("Content-Type", "text/html; charset=utf-8")
            self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            self.end_headers()
            self.wfile.write(html.encode())
        except FileNotFoundError:
            self.send_json({"code": 404, "msg": "HTML文件不存在"}, 404)

    def serve_qr_image(self):
        """提供缓存的QR码图片"""
        if self.__class__._last_qr_image:
            img_bytes = self.__class__._last_qr_image
            content_type = "image/jpeg"
            if img_bytes[:4] == b'\x89PNG':
                content_type = "image/png"
            self.send_cors(200)
            self.send_header("Content-Type", content_type)
            self.send_header("Content-Length", str(len(img_bytes)))
            self.send_header("Cache-Control", "no-cache, no-store")
            self.end_headers()
            self.wfile.write(img_bytes)
        else:
            self.send_error(404, "No QR code generated yet")

    def serve_static(self, path):
        """提供静态文件服务"""
        import mimetypes
        # 安全检查：防止路径遍历
        if ".." in path:
            self.send_error(403)
            return
        
        # 构建文件路径
        static_dir = os.path.dirname(HTML_PATH)
        file_path = os.path.join(static_dir, path.lstrip("/"))
        
        # 检查文件是否存在
        if not os.path.exists(file_path) or not os.path.isfile(file_path):
            self.send_error(404)
            return
        
        # 获取MIME类型
        mime_type, _ = mimetypes.guess_type(file_path)
        if mime_type is None:
            mime_type = "application/octet-stream"
        
        try:
            with open(file_path, "rb") as f:
                content = f.read()
            self.send_cors(200)
            self.send_header("Content-Type", mime_type)
            self.send_header("Content-Length", str(len(content)))
            # HTML永远不缓存（避免改代码后用户看到旧版）；其他静态资源缓存1小时
            if path.endswith(".html"):
                self.send_header("Cache-Control", "no-cache, no-store, must-revalidate")
            else:
                self.send_header("Cache-Control", "public, max-age=3600")
            self.end_headers()
            self.wfile.write(content)
        except Exception as e:
            self.send_error(500)

    def get_token(self):
        if os.path.exists(TOKEN_PATH):
            with open(TOKEN_PATH) as f:
                token = f.read().strip()
                if token:
                    return token
        return None

    def verify_token(self, token):
        """真实校验token是否有效（调用AMS接口验证）"""
        try:
            req = urllib.request.Request(
                f"{AMS_BASE}/hotel/web/basics/getHotelList",
                headers={"Authorization": f"Bearer {token}"}
            )
            with urllib.request.urlopen(req, timeout=8) as resp:
                result = json.loads(resp.read().decode())
                # AMS返回code 200=有效, 401=过期
                if isinstance(result, dict):
                    return result.get("code") == 200
                return False
        except urllib.error.HTTPError as e:
            return e.code != 401
        except Exception:
            # 网络异常时保守处理：认为token有效（避免误登出）
            return True

    def save_token(self, token):
        os.makedirs(os.path.dirname(TOKEN_PATH), exist_ok=True)
        with open(TOKEN_PATH, "w") as f:
            f.write(token)

    # ========== 账号权限系统 ==========

    def _auth_conn(self):
        conn = sqlite3.connect(AUTH_DB)
        conn.row_factory = sqlite3.Row
        return conn

    def _read_json_body(self):
        length = int(self.headers.get("Content-Length", 0))
        raw = self.rfile.read(length) if length else b""
        try:
            return json.loads(raw.decode("utf-8")) if raw else {}
        except Exception:
            return {}

    def _session_token(self):
        auth = self.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            return auth[7:].strip()
        return (self.headers.get("X-Session-Token") or "").strip()

    def _session_user(self):
        """返回当前会话用户（dict）；未登录/过期/停用返回 None"""
        token = self._session_token()
        if not token:
            return None
        try:
            conn = self._auth_conn()
            row = conn.execute("""SELECT u.* FROM sessions s JOIN users u ON u.id = s.user_id
                                  WHERE s.token = ? AND s.expires_at > datetime('now','localtime')
                                  AND u.enabled = 1""", (token,)).fetchone()
            conn.close()
            return dict(row) if row else None
        except Exception:
            return None

    def _new_session(self, user_id):
        token = secrets.token_urlsafe(32)
        conn = self._auth_conn()
        conn.execute("""INSERT INTO sessions (token, user_id, expires_at)
                        VALUES (?, ?, datetime('now','localtime','+%d days'))""" % SESSION_DAYS,
                     (token, user_id))
        conn.commit()
        conn.close()
        return token

    def _user_workspaces(self, user):
        ids = json.loads(user.get("workspace_ids") or "[]")
        if not ids:
            return []
        try:
            conn = self._auth_conn()
            q = ",".join("?" * len(ids))
            rows = conn.execute(f"SELECT * FROM workspaces WHERE id IN ({q}) ORDER BY id", ids).fetchall()
            conn.close()
            return [dict(r) for r in rows]
        except Exception:
            return []

    def _user_permissions(self, user):
        """返回用户权限：{modules:set, campuses:set}；管理员拥有全部模块、不限校区"""
        if not user:
            return {"modules": set(), "campuses": set()}
        if user.get("is_admin"):
            return {"modules": set(MODULES.keys()), "campuses": set()}
        modules, campuses = set(), set()
        for ws in self._user_workspaces(user):
            for m in json.loads(ws.get("modules") or "[]"):
                if m:
                    modules.add(m)
            for cid in json.loads(ws.get("campuses") or "[]"):
                if cid:
                    campuses.add(str(cid))
        return {"modules": modules, "campuses": campuses}

    def _campus_scope(self, user):
        """返回受限校区ID列表（str）；空列表表示不限校区"""
        if not user or user.get("is_admin"):
            return []
        return sorted(self._user_permissions(user)["campuses"])

    def _campus_scope_names(self, user):
        """返回受限校区的名称列表（布草等按校区名存储的表用）"""
        return [HOTEL_IDS.get(cid, cid) for cid in self._campus_scope(user)]

    def _campus_filter_sql(self, user, column="hotel_id"):
        """按用户校区范围生成 SQL 过滤片段；不限时返回 ("", [])"""
        scope = self._campus_scope(user)
        if not scope:
            return "", []
        ph = ",".join("?" * len(scope))
        return f" AND {column} IN ({ph})", list(scope)

    def _check_api_access(self, path):
        """模块权限校验。返回 (code, msg)；code 为 0 表示放行"""
        user = self._session_user()
        if not user:
            return 401, "未登录，请先登录"
        if user.get("is_admin"):
            return 0, None
        modules = self._user_permissions(user)["modules"]
        for prefix, mods in API_MODULE_RULES:
            if path.startswith(prefix):
                if any(m in modules for m in mods):
                    return 0, None
                return 403, "当前账号无权限访问该模块"
        return 0, None

    def _require_admin(self):
        user = self._session_user()
        if not user:
            return None, 401
        if not user.get("is_admin"):
            return None, 403
        return user, 0

    def auth_me(self):
        user = self._session_user()
        if not user:
            return self.send_json({"code": 401, "msg": "未登录"}, 401)
        perms = self._user_permissions(user)
        ws_list = self._user_workspaces(user)
        return self.send_json({
            "code": 200,
            "user": {"id": user["id"], "username": user["username"],
                     "display_name": user["display_name"], "is_admin": user["is_admin"]},
            "modules": sorted(perms["modules"]),
            "campuses": sorted(perms["campuses"]),
            "workspaces": [{"id": w["id"], "name": w["name"]} for w in ws_list],
            "all_modules": MODULES,
            "hotels": HOTEL_IDS,
        })

    def auth_login(self):
        data = self._read_json_body()
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""
        if not username or not password:
            return self.send_json({"code": 400, "msg": "缺少用户名或密码"}, 400)
        try:
            conn = self._auth_conn()
            row = conn.execute("SELECT * FROM users WHERE username = ?", (username,)).fetchone()
            conn.close()
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"登录失败: {e}"}, 500)
        if not row:
            return self.send_json({"code": 400, "msg": "账号不存在"}, 400)
        user = dict(row)
        if not user["enabled"]:
            return self.send_json({"code": 400, "msg": "账号已停用，请联系管理员"}, 400)
        if _hash_password(password, user["salt"]) != user["password_hash"]:
            return self.send_json({"code": 400, "msg": "密码错误"}, 400)
        token = self._new_session(user["id"])
        return self.send_json({
            "code": 200, "msg": "登录成功", "session_token": token,
            "user": {"id": user["id"], "username": user["username"],
                     "display_name": user["display_name"], "is_admin": user["is_admin"]},
        })

    def auth_logout(self):
        token = self._session_token()
        if token:
            try:
                conn = self._auth_conn()
                conn.execute("DELETE FROM sessions WHERE token = ?", (token,))
                conn.commit()
                conn.close()
            except Exception:
                pass
        return self.send_json({"code": 200, "msg": "已退出登录"})

    def auth_modules(self):
        user = self._session_user()
        if not user:
            return self.send_json({"code": 401, "msg": "未登录"}, 401)
        return self.send_json({"code": 200, "modules": MODULES, "hotels": HOTEL_IDS})

    # ---- 工作区管理（仅管理员写操作） ----

    def auth_workspaces_list(self):
        user = self._session_user()
        if not user:
            return self.send_json({"code": 401, "msg": "未登录"}, 401)
        try:
            conn = self._auth_conn()
            rows = conn.execute("SELECT * FROM workspaces ORDER BY id").fetchall()
            conn.close()
            return self.send_json({"code": 200, "workspaces": [dict(r) for r in rows]})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def auth_workspaces_create(self):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        data = self._read_json_body()
        name = (data.get("name") or "").strip()
        if not name:
            return self.send_json({"code": 400, "msg": "工作区名称必填"}, 400)
        modules = [m for m in (data.get("modules") or []) if m in MODULES]
        campuses = [str(c) for c in (data.get("campuses") or []) if c]
        try:
            conn = self._auth_conn()
            cur = conn.execute("INSERT INTO workspaces (name, modules, campuses, remark) VALUES (?,?,?,?)",
                               (name, json.dumps(modules, ensure_ascii=False),
                                json.dumps(campuses, ensure_ascii=False), data.get("remark", "")))
            conn.commit()
            new_id = cur.lastrowid
            conn.close()
            return self.send_json({"code": 200, "msg": "已创建", "id": new_id})
        except sqlite3.IntegrityError:
            return self.send_json({"code": 400, "msg": "工作区名称已存在"}, 400)
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def auth_workspaces_update(self, ws_id):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        data = self._read_json_body()
        name = (data.get("name") or "").strip()
        if not name:
            return self.send_json({"code": 400, "msg": "工作区名称必填"}, 400)
        modules = [m for m in (data.get("modules") or []) if m in MODULES]
        campuses = [str(c) for c in (data.get("campuses") or []) if c]
        try:
            conn = self._auth_conn()
            dup = conn.execute("SELECT id FROM workspaces WHERE name = ? AND id != ?", (name, ws_id)).fetchone()
            if dup:
                conn.close()
                return self.send_json({"code": 400, "msg": "工作区名称已存在"}, 400)
            conn.execute("UPDATE workspaces SET name=?, modules=?, campuses=?, remark=? WHERE id=?",
                         (name, json.dumps(modules, ensure_ascii=False),
                          json.dumps(campuses, ensure_ascii=False), data.get("remark", ""), ws_id))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "已保存"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def auth_workspaces_delete(self, ws_id):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        try:
            conn = self._auth_conn()
            conn.execute("DELETE FROM workspaces WHERE id = ?", (ws_id,))
            # 同步从所有账号的工作区列表中移除
            for u in conn.execute("SELECT id, workspace_ids FROM users").fetchall():
                ids = [i for i in json.loads(u["workspace_ids"] or "[]") if int(i) != int(ws_id)]
                conn.execute("UPDATE users SET workspace_ids=? WHERE id=?", (json.dumps(ids), u["id"]))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "已删除"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    # ---- 账号管理（仅管理员） ----

    def auth_users_list(self):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        try:
            conn = self._auth_conn()
            rows = conn.execute("SELECT * FROM users ORDER BY id").fetchall()
            conn.close()
            users = []
            for r in rows:
                d = dict(r)
                d.pop("password_hash", None)
                d.pop("salt", None)
                users.append(d)
            return self.send_json({"code": 200, "users": users})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def auth_users_create(self):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        data = self._read_json_body()
        username = (data.get("username") or "").strip()
        password = data.get("password") or ""
        if not username or not password:
            return self.send_json({"code": 400, "msg": "用户名和密码必填"}, 400)
        if len(password) < 4:
            return self.send_json({"code": 400, "msg": "密码至少4位"}, 400)
        salt = secrets.token_hex(16)
        try:
            conn = self._auth_conn()
            cur = conn.execute("""INSERT INTO users (username, password_hash, salt, display_name, workspace_ids, is_admin, enabled)
                                  VALUES (?,?,?,?,?,?,?)""",
                               (username, _hash_password(password, salt), salt,
                                data.get("display_name", ""),
                                json.dumps([int(i) for i in (data.get("workspace_ids") or []) if i]),
                                1 if data.get("is_admin") else 0,
                                1 if data.get("enabled", True) else 0))
            conn.commit()
            new_id = cur.lastrowid
            conn.close()
            return self.send_json({"code": 200, "msg": "已创建", "id": new_id})
        except sqlite3.IntegrityError:
            return self.send_json({"code": 400, "msg": "用户名已存在"}, 400)
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def auth_users_update(self, uid):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        data = self._read_json_body()
        try:
            conn = self._auth_conn()
            row = conn.execute("SELECT * FROM users WHERE id = ?", (uid,)).fetchone()
            if not row:
                conn.close()
                return self.send_json({"code": 400, "msg": "账号不存在"}, 400)
            user = dict(row)
            username = (data.get("username") or user["username"]).strip()
            dup = conn.execute("SELECT id FROM users WHERE username = ? AND id != ?", (username, uid)).fetchone()
            if dup:
                conn.close()
                return self.send_json({"code": 400, "msg": "用户名已存在"}, 400)
            conn.execute("UPDATE users SET username=?, display_name=?, workspace_ids=?, is_admin=?, enabled=? WHERE id=?",
                         (username, data.get("display_name", user["display_name"]),
                          json.dumps([int(i) for i in (data.get("workspace_ids") if data.get("workspace_ids") is not None else json.loads(user["workspace_ids"] or "[]")) if i]),
                          1 if data.get("is_admin", user["is_admin"]) else 0,
                          1 if data.get("enabled", user["enabled"]) else 0,
                          uid))
            password = data.get("password")
            if password:
                if len(password) < 4:
                    conn.close()
                    return self.send_json({"code": 400, "msg": "密码至少4位"}, 400)
                salt = secrets.token_hex(16)
                conn.execute("UPDATE users SET password_hash=?, salt=? WHERE id=?",
                             (_hash_password(password, salt), salt, uid))
                conn.execute("DELETE FROM sessions WHERE user_id = ?", (uid,))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "已保存"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def auth_users_delete(self, uid):
        _, code = self._require_admin()
        if code:
            return self.send_json({"code": code, "msg": "无权限，仅管理员可操作"}, code)
        try:
            conn = self._auth_conn()
            conn.execute("DELETE FROM users WHERE id = ?", (uid,))
            conn.execute("DELETE FROM sessions WHERE user_id = ?", (uid,))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "已删除"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def handle_login(self, data):
        """处理账号密码登录"""
        username = data.get("username")
        password = data.get("password")
        if not username or not password:
            self.send_json({"code": 400, "msg": "缺少用户名或密码"}, 400)
            return

        # 先获取验证码
        try:
            req = urllib.request.Request(f"{AMS_BASE}/hotel/web/captchaImage")
            with urllib.request.urlopen(req, timeout=10) as resp:
                captcha_data = json.loads(resp.read().decode())
        except Exception as e:
            self.send_json({"code": 500, "msg": f"获取验证码失败: {str(e)}"}, 500)
            return

        # 登录
        login_data = json.dumps({
            "username": username,
            "password": password,
            "code": data.get("code", ""),
            "uuid": captcha_data.get("uuid", "")
        }).encode()
        req = urllib.request.Request(
            f"{AMS_BASE}/hotel/web/login",
            data=login_data,
            headers={"Content-Type": "application/json"}
        )
        try:
            with urllib.request.urlopen(req, timeout=10) as resp:
                result = json.loads(resp.read().decode())
                if result.get("token"):
                    self.save_token(result["token"])
                    # AMS 登录成功后按身份匹配本地账号并建立会话
                    session_token, account_name, matched, is_new = self._session_for_ams_identity()
                    self.send_json({"code": 200, "msg": "登录成功",
                                    "session_token": session_token, "account_name": account_name,
                                    "matched": matched, "is_new": is_new})
                else:
                    self.send_json({"code": 400, "msg": result.get("msg", "登录失败")}, 400)
        except Exception as e:
            self.send_json({"code": 500, "msg": f"登录失败: {str(e)}"}, 500)

    def _admin_session(self):
        """为管理员账号建立/复用会话（AMS登录或扫码登录成功后调用）"""
        try:
            conn = self._auth_conn()
            row = conn.execute("SELECT id FROM users WHERE is_admin=1 AND enabled=1 ORDER BY id LIMIT 1").fetchone()
            conn.close()
            if row:
                return self._new_session(row["id"])
        except Exception:
            pass
        return None

    def _session_for_ams_identity(self):
        """扫码/AMS密码登录后，把AMS身份映射到本地账号：
        1) 用刚保存的AMS token 调 getInfo 获取扫码人姓名
        2) 按 本地用户名 或 显示姓名 精确匹配启用的账号 → 为该账号建会话
        3) 未匹配时自动创建本地账号（默认绑定第一个非管理后台工作区，密码随机，仅扫码可登录）
        返回 (session_token, 账号显示名或None, 是否匹配/注册成功, 是否新注册)
        """
        token = self.get_token()
        identity = None
        if token:
            try:
                req = urllib.request.Request(f"{AMS_BASE}/hotel/web/getInfo",
                                             headers={"Authorization": f"Bearer {token}"})
                with urllib.request.urlopen(req, timeout=8) as resp:
                    data = json.loads(resp.read().decode())
                user = (data or {}).get("user") or {}
                identity = (user.get("userName") or user.get("nickName") or "").strip()
            except Exception:
                identity = None
        if identity:
            try:
                conn = self._auth_conn()
                row = conn.execute(
                    "SELECT id, display_name, username FROM users "
                    "WHERE enabled=1 AND (username=? OR display_name=?) "
                    "ORDER BY is_admin ASC, id ASC LIMIT 1", (identity, identity)).fetchone()
                if row:
                    conn.close()
                    return self._new_session(row["id"]), row["display_name"] or row["username"], True, False
                # 未匹配：自动创建本地账号（若存在同名禁用账号则启用）
                existing = conn.execute("SELECT id, display_name, username FROM users WHERE username=? ORDER BY id LIMIT 1",
                                        (identity,)).fetchone()
                if existing:
                    conn.execute("UPDATE users SET display_name=?, enabled=1 WHERE id=?",
                                 (identity, existing["id"]))
                    user_id = existing["id"]
                    conn.commit()
                    conn.close()
                    return self._new_session(user_id), identity, True, True
                ws = conn.execute("SELECT id FROM workspaces WHERE name != '管理后台' ORDER BY id LIMIT 1").fetchone()
                ws_ids = json.dumps([ws["id"]]) if ws else "[]"
                salt = secrets.token_hex(16)
                cur = conn.execute(
                    "INSERT INTO users (username, password_hash, salt, display_name, workspace_ids, is_admin, enabled) "
                    "VALUES (?,?,?,?,?,0,1)",
                    (identity, _hash_password(secrets.token_urlsafe(12), salt), salt, identity, ws_ids))
                user_id = cur.lastrowid
                conn.commit()
                conn.close()
                return self._new_session(user_id), identity, True, True
            except Exception:
                pass
        return self._admin_session(), identity, False, False

    def auth_me_update(self):
        """当前账号完善/修改本人资料（姓名）"""
        user = self._session_user()
        if not user:
            return self.send_json({"code": 401, "msg": "未登录"}, 401)
        data = self._read_json_body()
        display_name = (data.get("display_name") or "").strip()
        if not display_name:
            return self.send_json({"code": 400, "msg": "姓名不能为空"}, 400)
        try:
            conn = self._auth_conn()
            conn.execute("UPDATE users SET display_name=? WHERE id=?", (display_name, user["id"]))
            conn.commit()
            conn.close()
            return self.send_json({"code": 200, "msg": "已保存"})
        except Exception as e:
            return self.send_json({"code": 500, "msg": str(e)}, 500)

    def proxy_api(self, api_path):
        token = self.get_token()
        # 登录相关公开接口无需token（验证码/登录）
        public_prefixes = ("/hotel/web/captchaImage", "/hotel/web/login")
        if not token and not api_path.startswith(public_prefixes):
            return self.send_json({"code": 401, "msg": "未登录，请先扫码登录"}, 401)

        from urllib.parse import urlparse
        parsed = urlparse(api_path)
        
        # 所有请求都走AMS主域名
        url = f"{AMS_BASE}{parsed.path}"
        
        if parsed.query:
            url += "?" + parsed.query

        headers = {"Content-Type": "application/json"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        req = urllib.request.Request(url, headers=headers)
        try:
            with urllib.request.urlopen(req, timeout=15) as resp:
                self.send_json(json.loads(resp.read().decode()))
        except urllib.error.HTTPError as e:
            error_body = e.read().decode()
            try:
                error_json = json.loads(error_body)
                self.send_json(error_json, e.code)
            except:
                self.send_json({"code": e.code, "msg": error_body}, e.code)
        except Exception as e:
            self.send_json({"code": 500, "msg": str(e)}, 500)

    def qr_login(self):
        """生成微信扫码登录二维码"""
        u = uuid.uuid4().hex
        try:
            req = urllib.request.Request(f"{AMS_BASE}/hotel/web/system/qrcode/generate/{u}")
            with urllib.request.urlopen(req, timeout=10) as resp:
                result = json.loads(resp.read().decode())
                # 缓存图片供/qr-image使用
                if result.get("data", {}).get("imageUrl"):
                    img_data = result["data"]["imageUrl"]
                    if img_data.startswith("data:image"):
                        import base64
                        b64_data = img_data.split(",")[1]
                        self.__class__._last_qr_image = base64.b64decode(b64_data)
                        self.__class__._last_qr_token = result["data"].get("token") or result["data"].get("scene")
                self.send_json(result)
        except Exception as e:
            self.send_json({"code": 500, "msg": f"生成二维码失败: {str(e)}"}, 500)

    _last_qr_image = None
    _last_qr_token = None

    def qr_status(self):
        """查询扫码状态"""
        from urllib.parse import parse_qs, urlparse
        params = parse_qs(urlparse(self.path).query)
        uuid = params.get("uuid", [None])[0]
        if not uuid:
            return self.send_json({"code": 400, "msg": "缺少uuid参数"}, 400)

        try:
            req = urllib.request.Request(f"{AMS_BASE}/hotel/web/system/qrcode/status/{uuid}")
            with urllib.request.urlopen(req, timeout=10) as resp:
                self.send_json(json.loads(resp.read().decode()))
        except Exception as e:
            self.send_json({"code": 500, "msg": str(e)}, 500)

    def qr_auth(self, uuid):
        """扫码登录认证"""
        try:
            req = urllib.request.Request(f"{AMS_BASE}/hotel/web/system/qrcode/auth/login/{uuid}")
            with urllib.request.urlopen(req, timeout=10) as resp:
                result = json.loads(resp.read().decode())
                if result.get("token"):
                    self.save_token(result["token"])
                    # 扫码成功后按身份匹配本地账号并建立会话
                    result["session_token"], result["account_name"], result["matched"], result["is_new"] = self._session_for_ams_identity()
                self.send_json(result)
        except Exception as e:
            self.send_json({"code": 500, "msg": str(e)}, 500)

    def log_message(self, format, *args):
        """自定义日志格式：记录时间/客户端/请求路径/状态码"""
        try:
            from urllib.parse import urlparse
            path = urlparse(self.path).path
            # 静默健康检查请求，避免刷屏
            if path in ("/api/status", "/favicon.ico"):
                return
            # args来自log_request: (requestline, code, size)
            code = args[1] if len(args) > 1 else '-'
            print(f"[{self.log_date_time_string()}] {self.client_address[0]} {self.command} {path} {code}")
        except Exception:
            print(f"[{self.log_date_time_string()}] {args[0] if args else ''}")

    def do_HEAD(self):
        """支持HEAD请求（诊断用）"""
        from urllib.parse import urlparse
        parsed_path = urlparse(self.path)
        path = parsed_path.path
        if path.endswith((".css", ".js", ".png", ".jpg", ".gif", ".ico", ".svg", ".woff", ".woff2", ".ttf", ".html")):
            self.serve_static(path)
        else:
            self.send_cors(200)
            self.end_headers()

    # ========== 工资表 ==========

    # 员工日薪配置（2026-08-03 李瑞峰确认）
    PAYROLL_CONFIG = {
        # 日薪（元/天）2026-08-03 李瑞峰确认
        "day_rate": {
            "柴浩凯": 150,   # 公寓一组组长
            "杨苏元": 130,   # 酒店前台组长
            "吴国飞": 130,   # 公寓二组组长
            # 库管4人：考勤表职位未体现，按130算
            "丁一博": 130,
            "袁术飞": 130,
            "陈韬": 130,
            "李瑞峰": 130,
        },
        "default_rate": 110,  # 生活班主任
        "front_desk_rate": 130,  # 酒店前台
        # 组长补助（元/月）
        "subsidy": {
            "柴浩凯": 800,
            "杨苏元": 500,
            "吴国飞": 500,
        },
        # 固定底薪
        "base_salary": 2000,
        "manager_salary": 7500,  # 王怔鑫主管
        # 前台职位关键词
        "front_desk_keywords": ["前台"],
        "manager_name": "王怔鑫",
    }

    def handle_payroll_parse(self):
        """解析考勤表Excel，返回员工出勤数据"""
        import base64, io
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body.decode())
            b64 = data.get("file_b64", "")
            if not b64:
                return self.send_json({"code": 400, "msg": "缺少文件数据"}, 400)
            file_bytes = base64.b64decode(b64)
            # 解析Excel
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
            ws = wb['月度汇总'] if '月度汇总' in wb.sheetnames else wb[wb.sheetnames[0]]

            employees = []
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row[0]:
                    continue
                name = str(row[0]).strip()
                group = str(row[1] or '').strip()
                position = str(row[4] or '').strip()
                should = row[7]
                actual = row[8]
                work_hours = row[9]
                late = row[10]
                early = row[11]
                absent = row[17]
                leave = row[19]

                # 判断日薪
                rate = self.PAYROLL_CONFIG["day_rate"].get(name)
                if rate is None:
                    if any(kw in position for kw in self.PAYROLL_CONFIG["front_desk_keywords"]):
                        rate = self.PAYROLL_CONFIG["front_desk_rate"]
                    else:
                        rate = self.PAYROLL_CONFIG["default_rate"]

                # 底薪
                base = self.PAYROLL_CONFIG["manager_salary"] if name == self.PAYROLL_CONFIG["manager_name"] else self.PAYROLL_CONFIG["base_salary"]
                # 组长补助
                subsidy = self.PAYROLL_CONFIG["subsidy"].get(name, 0)

                # 当月天数 = 考勤表列数（col44-74是日期列，数非空日期）
                # 简化：从考勤表的日期列（44-74）推断当月天数
                month_days = 31
                try:
                    date_cols = [row[i] for i in range(43, 74) if row[i] is not None]
                    # 数唯一数字日期（排除'六'/'日'）
                    day_nums = [v for v in date_cols if isinstance(v, (int, float))]
                    if day_nums:
                        month_days = int(max(day_nums))
                except:
                    pass

                # 请假合计（col20-43，索引19-42）
                leave_total = 0
                try:
                    for i in range(19, 43):
                        v = row[i]
                        if v is not None:
                            fv = float(v)
                            if fv != 0:
                                leave_total += fv
                except:
                    pass

                # 在岗天数 = 当月天数 - 请假
                attend_days = month_days - leave_total

                # 王怔鑫特殊：默认在岗工资980、绩效1000（考核项，每月可改）
                if name == self.PAYROLL_CONFIG["manager_name"]:
                    employees.append({
                        "name": name,
                        "group": group,
                        "position": position,
                        "base_salary": base,
                        "attend_days": 0,
                        "leave_days": 0,
                        "day_rate": 0,
                        "subsidy": 0,
                        "on_duty_pay": 2480,  # 标准化1500+入住率980（手填）
                        "performance": 1000,   # 学生满意度（手填）
                        "deduction": 0,
                        "overtime": 0,
                    })
                    continue

                employees.append({
                    "name": name,
                    "group": group,
                    "position": position,
                    "base_salary": base,
                    "attend_days": round(attend_days, 1),
                    "leave_days": round(leave_total, 1),
                    "day_rate": rate,
                    "subsidy": subsidy,
                    "on_duty_pay": round(attend_days * rate, 2),
                    "performance": 0,  # 绩效手动填
                    "deduction": 0,     # 扣款手动填
                    "overtime": 0,      # 加班手动填
                })

            return self.send_json({"code": 200, "employees": employees, "count": len(employees)})
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"解析失败: {str(e)}"}, 500)

    def handle_payroll_export(self):
        """基于7月模板生成工资表Excel（酒店发+新途径发，格式100%保留）"""
        import base64, io, os
        try:
            content_length = int(self.headers.get('Content-Length', 0))
            body = self.rfile.read(content_length)
            data = json.loads(body.decode())
            employees = data.get("employees", [])
            export_type = data.get("export_type", "all")  # hotel / new / all
            # 绩效：酒店公司/新途径公司统一（前端传入）
            perf_hotel = float(data.get("perf_hotel", 420) or 420)
            perf_new = float(data.get("perf_new", 420) or 420)

            # 模板路径
            tpl_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "payroll_templates")
            hotel_tpl = os.path.join(tpl_dir, "酒店发模板.xlsx")
            new_tpl = os.path.join(tpl_dir, "新途径发模板.xlsx")

            if not os.path.exists(hotel_tpl) or not os.path.exists(new_tpl):
                return self.send_json({"code": 500, "msg": "模板文件不存在"}, 500)

            # 员工数据索引
            emp_map = {e["name"]: e for e in employees}

            # 模板行位配置：{name: (row, is_hotel)}
            hotel_rows = {
                "李静怡": 4, "丁一博": 5, "袁术飞": 6, "陈韬": 7, "贾敏": 8,
                "于洋": 9, "李瑞峰": 10, "李文臻": 11, "云欣悦": 12, "崔佳乐": 13,
                "高博": 14, "杨苏元": 15, "柴浩凯": 18, "王怔鑫": 21,
            }
            new_rows = {
                "史琴": 4, "杨澜": 5, "常洪瑄": 6, "姜亚鑫": 7, "景鹏飞": 8,
                "呼斯乐": 9, "闫瑾": 10, "吴国飞": 12,
            }

            import openpyxl

            def update_sheet(tpl_path, row_map, is_hotel):
                wb = openpyxl.load_workbook(tpl_path)
                ws = wb[wb.sheetnames[0]]
                for name, row in row_map.items():
                    e = emp_map.get(name)
                    if not e:
                        continue
                    # 王怔鑫特殊（第3段表头）：col5姓名 col6底薪7500 col7标准化1500 col8入住率 col9满意度 col12应发
                    if name == "王怔鑫":
                        base = 7500
                        on_duty = e.get("on_duty_pay", 2480) or 2480  # 标准化1500+入住率980
                        perf = e.get("performance", 1000) or 1000
                        std = 1500          # 标准化（固定1500）
                        occ_rate = round(on_duty - std, 2)  # 入住率 = 合计-1500
                        total = base + std + occ_rate + perf
                        ws.cell(row, 6, base)
                        ws.cell(row, 7, std)
                        ws.cell(row, 8, occ_rate)
                        ws.cell(row, 9, perf)
                        ws.cell(row, 12, total)
                        continue
                    # 普通员工/组长
                    base = e.get("base_salary", 2000)
                    days = e.get("attend_days", 0)
                    rate = e.get("day_rate", 110)
                    on_duty = round(days * rate, 2)
                    subsidy = e.get("subsidy", 0) or 0
                    deduction = e.get("deduction", 0) or 0
                    overtime = e.get("overtime", 0) or 0
                    # 绩效按公司统一
                    perf = perf_hotel if is_hotel else perf_new
                    total = base + on_duty + perf + subsidy - deduction + overtime
                    if is_hotel:
                        # 酒店发：姓名col5 底薪col6 天数col7 在岗col8 绩效col9 补助col11 应发col12
                        ws.cell(row, 6, base)
                        ws.cell(row, 7, days)
                        ws.cell(row, 8, on_duty)
                        ws.cell(row, 9, perf)
                        ws.cell(row, 11, subsidy)
                        ws.cell(row, 12, total)
                    else:
                        # 新途径发：姓名col4 底薪col5 天数col6 在岗col7 绩效col8 补助col10 应发col11
                        ws.cell(row, 5, base)
                        ws.cell(row, 6, days)
                        ws.cell(row, 7, on_duty)
                        ws.cell(row, 8, perf)
                        ws.cell(row, 10, subsidy)
                        ws.cell(row, 11, total)
                return wb

            # 生成两张表
            wb_hotel = update_sheet(hotel_tpl, hotel_rows, True)
            wb_new = update_sheet(new_tpl, new_rows, False)

            def save_wb(wb_src, title):
                out = openpyxl.Workbook()
                del out[out.sheetnames[0]]
                ws_src = wb_src[wb_src.sheetnames[0]]
                ws_new = out.create_sheet(title)
                for row in ws_src.iter_rows():
                    for cell in row:
                        nc = ws_new.cell(cell.row, cell.column, cell.value)
                        if cell.has_style:
                            nc.font = cell.font.copy()
                            nc.border = cell.border.copy()
                            nc.fill = cell.fill.copy()
                            nc.alignment = cell.alignment.copy()
                            nc.number_format = cell.number_format
                for col_letter, dim in ws_src.column_dimensions.items():
                    if dim.width:
                        ws_new.column_dimensions[col_letter].width = dim.width
                for mc in ws_src.merged_cells.ranges:
                    ws_new.merge_cells(str(mc))
                for r, dim in ws_src.row_dimensions.items():
                    if dim.height:
                        ws_new.row_dimensions[r].height = dim.height
                buf = io.BytesIO()
                out.save(buf)
                return base64.b64encode(buf.getvalue()).decode()

            # 按导出类型返回（文件名用当前月份）
            from datetime import datetime
            month_str = datetime.now().strftime("%m月")
            if export_type == "hotel":
                return self.send_json({"code": 200, "file_b64": save_wb(wb_hotel, '酒店发'), "filename": f"上岸公寓{month_str}工资表（酒店发）.xlsx"})
            elif export_type == "new":
                return self.send_json({"code": 200, "file_b64": save_wb(wb_new, '新途径发'), "filename": f"上岸公寓{month_str}工资表（新途径发）.xlsx"})
            else:
                return self.send_json({"code": 200, "files": [
                    {"file_b64": save_wb(wb_hotel, '酒店发'), "filename": f"上岸公寓{month_str}工资表（酒店发）.xlsx"},
                    {"file_b64": save_wb(wb_new, '新途径发'), "filename": f"上岸公寓{month_str}工资表（新途径发）.xlsx"}
                ]})
        except Exception as e:
            return self.send_json({"code": 500, "msg": f"导出失败: {str(e)}"}, 500)

if __name__ == "__main__":
    init_auth_db()
    print(f"AMS代理服务器启动在端口 {PORT}")
    print(f"访问 http://localhost:{PORT} 或 http://192.168.31.179:{PORT}")
    server = ThreadedHTTPServer(("0.0.0.0", PORT), ProxyHandler)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n服务器已停止")
        server.shutdown()
